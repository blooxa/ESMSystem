from rest_framework import viewsets, permissions, status, filters
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.authtoken.models import Token
from django.contrib.auth.models import User
from django.contrib.auth.hashers import make_password, check_password
from django.utils import timezone
from django.db.models.signals import post_save
from django.dispatch import receiver
from asgiref.sync import async_to_sync
from channels.layers import get_channel_layer
from datetime import date, timedelta, datetime
from dateutil.relativedelta import relativedelta  # Добавьте эту строку
from .models import (
    Request, RequestHistory, AppUser, Shop, Employee, Position,
    Nomenclature, IssuanceStandard, PPEIssue,
    RequestEmployee, RequestEmployeeItem,
    PPEIssueStandard, PPEIssueRecord
)
from datetime import date, timedelta, datetime
from django.db import transaction
from io import BytesIO
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from .serializers import (
    RequestSerializer, RequestCreateSerializer,
    RequestApproveSerializer, RequestOrderSerializer,
    UserSerializer, RequestHistorySerializer,
    NomenclatureSerializer
)
# В начале файла добавьте импорт
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from rest_framework.authtoken.models import Token
from django.contrib.auth.hashers import check_password
from .models import AppUser
from django.contrib.auth.models import User as DjangoUser

@api_view(['POST'])
@permission_classes([AllowAny])
def simple_auth(request):
    username = request.data.get('username')
    password = request.data.get('password')

    if not username or not password:
        return Response({'error': 'Логин и пароль обязательны'}, status=400)

    try:
        app_user = AppUser.objects.filter(login=username).first()
        if not app_user:
            return Response({'error': 'Неверный логин или пароль'}, status=401)

        if check_password(password, app_user.password):
            django_user, created = User.objects.get_or_create(
                username=app_user.login,
                defaults={'email': f"{app_user.login}@example.com", 'is_active': True}
            )

            token, _ = Token.objects.get_or_create(user=django_user)

            return Response({
                'token': token.key,
                'username': django_user.username,
                'is_superuser': django_user.is_superuser,
                'user_id': django_user.id,
                'role': app_user.role
            })
        else:
            return Response({'error': 'Неверный логин или пароль'}, status=401)
    except Exception as e:
        print(f"Auth error: {str(e)}")
        return Response({'error': 'Внутренняя ошибка сервера'}, status=500)

# ==================== PERMISSIONS ====================

class IsDepartmentHead(permissions.BasePermission):
    """Проверка, является ли пользователь начальником отдела"""
    def has_permission(self, request, view):
        return request.user.is_authenticated and (
                request.user.groups.filter(name='department_head').exists() or
                request.user.is_superuser
        )

class IsEconomicHead(permissions.BasePermission):
    """Проверка, является ли пользователь начальником хоз. отдела"""
    def has_permission(self, request, view):
        if not request.user.is_authenticated:
            return False
        try:
            app_user = AppUser.objects.get(login=request.user.username)
            return app_user.role == 'economic_head' or request.user.is_superuser
        except AppUser.DoesNotExist:
            return request.user.is_superuser

class IsAdminOrEconomicHead(permissions.BasePermission):
    """Проверка, является ли пользователь администратором или начальником хоз. отдела"""
    def has_permission(self, request, view):
        return request.user.is_authenticated and (
                request.user.is_superuser or
                request.user.groups.filter(name='economic_head').exists()
        )


class MassIssueViewSet(viewsets.ViewSet):
    """API для массовой выдачи СИЗ"""
    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=['post'])
    def mass_issue(self, request):
        """Массовая выдача СИЗ сотрудникам по нормам"""
        try:
            app_user = AppUser.objects.get(login=request.user.username)

            # Проверяем права (только admin и economic_head)
            if not (request.user.is_superuser or app_user.role in ['admin', 'economic_head', 'department_head']):
                return Response({'error': 'Доступ запрещен'}, status=403)

            data = request.data
            employees_data = data.get('employees', [])
            issue_date = data.get('issue_date', date.today().isoformat())

            if not employees_data:
                return Response({'error': 'Выберите хотя бы одного сотрудника'}, status=400)

            results = []
            errors = []

            for emp_data in employees_data:
                employee_id = emp_data.get('employee_id')
                items = emp_data.get('items', [])

                try:
                    employee = Employee.objects.get(employee_id=employee_id)
                except Employee.DoesNotExist:
                    errors.append(f'Сотрудник с ID {employee_id} не найден')
                    continue

                for item in items:
                    nomenclature_id = item.get('nomenclature_id')
                    size = item.get('selected_size', '')
                    quantity = item.get('selected_quantity', 0)
                    period_months = item.get('period_months', 12)
                    comment = item.get('comment', '')

                    if quantity <= 0:
                        continue

                    try:
                        nomenclature = Nomenclature.objects.get(nomenclature_id=nomenclature_id)
                    except Nomenclature.DoesNotExist:
                        errors.append(f'СИЗ с ID {nomenclature_id} не найден')
                        continue

                    # Рассчитываем дату следующей выдачи
                    from dateutil.relativedelta import relativedelta
                    issue_date_obj = datetime.strptime(issue_date, '%Y-%m-%d').date()
                    next_issue_date = issue_date_obj + relativedelta(months=period_months)

                    # Создаем запись о выдаче
                    issue = PPEIssueRecord.objects.create(
                        employee=employee,
                        nomenclature=nomenclature,
                        issue_date=issue_date_obj,
                        size=size,
                        quantity=quantity,
                        next_issue_date=next_issue_date,
                        issued_by=app_user,
                        comment=comment
                    )

                    results.append({
                        'employee_name': employee.full_name,
                        'nomenclature_title': nomenclature.title,
                        'quantity': quantity,
                        'next_issue_date': next_issue_date.isoformat()
                    })

            return Response({
                'message': f'Успешно выдано {len(results)} позиций',
                'results': results,
                'errors': errors
            }, status=201)

        except AppUser.DoesNotExist:
            return Response({'error': 'User not found'}, status=404)
        except Exception as e:
            return Response({'error': str(e)}, status=400)

    @action(detail=False, methods=['get'])
    def get_employee_standards(self, request):
        """Получить нормы выдачи для сотрудника (для массовой выдачи)"""
        employee_id = request.query_params.get('employee_id')
        if not employee_id:
            return Response({'error': 'employee_id required'}, status=400)

        try:
            employee = Employee.objects.get(employee_id=employee_id)
        except Employee.DoesNotExist:
            return Response({'error': 'Employee not found'}, status=404)

        # Получаем нормы выдачи для должности сотрудника
        standards = PPEIssueStandard.objects.filter(
            position=employee.position,
            is_active=True
        ).select_related('nomenclature')

        today = date.today()
        result = []

        for standard in standards:
            # Проверяем последнюю выдачу
            last_issue = PPEIssueRecord.objects.filter(
                employee=employee,
                nomenclature=standard.nomenclature
            ).order_by('-issue_date').first()

            can_issue = True
            last_issue_date = None
            recommended_size = None

            if last_issue:
                last_issue_date = last_issue.issue_date
                # Если срок еще не истек, но все равно можно выдать, просто предупредим
                if last_issue.next_issue_date and last_issue.next_issue_date > today:
                    can_issue = True  # Разрешаем выдачу даже если срок не истек

            # Определяем рекомендуемый размер
            title_lower = standard.nomenclature.title.lower()
            if 'обув' in title_lower or 'сапог' in title_lower or 'ботин' in title_lower:
                if employee.shoesize:
                    recommended_size = str(employee.shoesize)
                else:
                    recommended_size = '42' if employee.gender == 'M' else '38'
            elif 'каск' in title_lower or 'шлем' in title_lower or 'шапк' in title_lower:
                if employee.headsize:
                    recommended_size = str(employee.headsize)
                else:
                    recommended_size = '58'
            else:
                if employee.clothing_size:
                    recommended_size = employee.clothing_size
                elif employee.heightcm:
                    if employee.heightcm < 166:
                        recommended_size = '88'
                    elif employee.heightcm < 174:
                        recommended_size = '96'
                    elif employee.heightcm < 182:
                        recommended_size = '104'
                    else:
                        recommended_size = '112'
                else:
                    recommended_size = '100'

            result.append({
                'nomenclature_id': standard.nomenclature.nomenclature_id,
                'nomenclature_title': standard.nomenclature.title,
                'unit': standard.nomenclature.unit,
                'standard_quantity': float(standard.quantity),
                'period_months': standard.period_months,
                'last_issue_date': last_issue_date,
                'can_issue': can_issue,
                'recommended_size': recommended_size,
                'selected_size': recommended_size,
                'selected_quantity': float(standard.quantity),
            })

        return Response(result)
class IsAdminUserPermission(permissions.BasePermission):
    """Проверка, является ли пользователь администратором"""
    def has_permission(self, request, view):
        return request.user.is_authenticated and request.user.is_superuser

class IsSafetyOfficer(permissions.BasePermission):
    """Проверка, является ли пользователь сотрудником охраны труда"""
    def has_permission(self, request, view):
        if not request.user.is_authenticated:
            return False
        try:
            app_user = AppUser.objects.get(login=request.user.username)
            return app_user.role == 'safety_officer' or request.user.is_superuser
        except AppUser.DoesNotExist:
            return request.user.is_superuser

class IsSafetyOrAdmin(permissions.BasePermission):
    """Проверка, является ли пользователь сотрудником охраны труда или администратором"""
    def has_permission(self, request, view):
        return request.user.is_authenticated and (
            request.user.is_superuser or
            request.user.groups.filter(name='safety_officer').exists()
        )
class IsDepartmentHead(permissions.BasePermission):
    """Проверка, является ли пользователь начальником отдела"""
    def has_permission(self, request, view):
        return request.user.is_authenticated and (
                request.user.groups.filter(name='department_head').exists() or
                request.user.is_superuser
        )

class IsEconomicHead(permissions.BasePermission):
    """Проверка, является ли пользователь начальником хоз. отдела"""
    def has_permission(self, request, view):
        if not request.user.is_authenticated:
            return False
        try:
            app_user = AppUser.objects.get(login=request.user.username)
            return app_user.role == 'economic_head' or request.user.is_superuser
        except AppUser.DoesNotExist:
            return request.user.is_superuser
class IsAdminOrEconomicHead(permissions.BasePermission):
    """Проверка, является ли пользователь администратором или начальником хоз. отдела"""
    def has_permission(self, request, view):
        return request.user.is_authenticated and (
                request.user.is_superuser or
                request.user.groups.filter(name='economic_head').exists()
        )
class IsAdminUserPermission(permissions.BasePermission):
    """Проверка, является ли пользователь администратором"""
    def has_permission(self, request, view):
        return request.user.is_authenticated and request.user.is_superuser


# ==================== WEBSOCKET NOTIFICATIONS ====================

# ВРЕМЕННО ОТКЛЮЧЕНО - WebSocket не настроен на Render
def send_websocket_notification(request_obj, action):
    """Отправка уведомления через WebSocket (временно отключено)"""
    pass
    # channel_layer = get_channel_layer()
    #
    # request_data = {
    #     'id': request_obj.request_id,
    #     'title': request_obj.title,
    #     'status': request_obj.get_status_display(),
    #     'status_code': request_obj.status,
    #     'action': action,
    #     'requester_name': request_obj.user.employee.full_name if request_obj.user and request_obj.user.employee else request_obj.user.login,
    #     'updated_at': request_obj.updated_at.isoformat()
    # }
    #
    # async_to_sync(channel_layer.group_send)(
    #     'economic_head_group',
    #     {
    #         'type': 'request_update',
    #         'request': request_data
    #     }
    # )
    #
    # async_to_sync(channel_layer.group_send)(
    #     f'user_{request_obj.user.user_id}',
    #     {
    #         'type': 'request_update',
    #         'request': request_data
    #     }
    # )


@receiver(post_save, sender=Request)
def request_post_save(sender, instance, created, **kwargs):
    # Временно отключено
    pass
    # if created:
    #     send_websocket_notification(instance, 'created')
    # else:
    #     send_websocket_notification(instance, 'updated')


# ==================== REQUEST VIEWS ====================

class RequestViewSet(viewsets.ModelViewSet):
    serializer_class = RequestSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.OrderingFilter, filters.SearchFilter]
    search_fields = ['title', 'description', 'supplier_name']
    ordering_fields = ['created_at', 'status', 'updated_at']
    ordering = ['-created_at']

    @action(detail=True, methods=['post'], permission_classes=[IsEconomicHead])
    def mark_delivered(self, request, pk=None):
        """Отметить поступление СИЗ на склад"""
        request_obj = self.get_object()

        if request_obj.status not in ['ordered', 'partially_delivered']:
            return Response({'error': 'Можно отметить доставку только для заказанных заявок'}, status=400)

        delivered_items = request.data.get('delivered_items', [])
        comment = request.data.get('comment', '')

        if not delivered_items:
            return Response({'error': 'Укажите какие СИЗ поступили'}, status=400)

        # Проверяем, все ли СИЗ доставлены
        request_employees = RequestEmployee.objects.filter(request=request_obj)
        total_items = 0
        delivered_count = 0

        for req_emp in request_employees:
            items = RequestEmployeeItem.objects.filter(request_employee=req_emp)
            total_items += items.count()
            for item in items:
                for delivered in delivered_items:
                    if (delivered.get('employee_id') == req_emp.employee.employee_id and
                            delivered.get('nomenclature_id') == item.nomenclature.nomenclature_id):
                        delivered_count += 1
                        break

        if delivered_count >= total_items:
            request_obj.status = 'delivered'
        else:
            request_obj.status = 'partially_delivered'

        request_obj.comment = comment
        request_obj.save()

        RequestHistory.objects.create(
            request=request_obj,
            status_from=request_obj.status,
            status_to=request_obj.status,
            changed_by=request.user,
            comment=f'Отмечено поступление: {comment}'
        )

        return Response(RequestSerializer(request_obj).data)

    @action(detail=True, methods=['put'], permission_classes=[IsAuthenticated])
    def update_full_request(self, request, pk=None):
        """Полное обновление заявки (сотрудники и СИЗ)"""
        request_obj = self.get_object()

        # Проверяем, что заявка в статусе pending
        if request_obj.status != 'pending':
            return Response(
                {'error': 'Можно редактировать только заявки в статусе "На рассмотрении"'},
                status=400
            )

        # Проверяем права
        try:
            app_user = AppUser.objects.get(login=request.user.username)
            if request_obj.user != app_user and not request.user.is_superuser:
                return Response({'error': 'Нет прав для редактирования этой заявки'}, status=403)
        except AppUser.DoesNotExist:
            return Response({'error': 'User not found'}, status=404)

        data = request.data
        title = data.get('title')
        employees_data = data.get('employees', [])

        if not title:
            return Response({'error': 'Title required'}, status=400)
        if not employees_data:
            return Response({'error': 'At least one employee required'}, status=400)

        # Обновляем заголовок и описание
        request_obj.title = title
        request_obj.description = data.get('description', '')
        request_obj.save()

        # Удаляем существующих сотрудников и их СИЗ
        RequestEmployee.objects.filter(request=request_obj).delete()

        # Создаем заново
        for emp_data in employees_data:
            employee = Employee.objects.get(employee_id=emp_data['employee_id'])
            request_employee = RequestEmployee.objects.create(
                request=request_obj,
                employee=employee,
                height=emp_data.get('height')
            )
            for item in emp_data['items']:
                if item.get('selected_quantity', 0) > 0:
                    nomenclature = Nomenclature.objects.get(nomenclature_id=item['nomenclature_id'])
                    RequestEmployeeItem.objects.create(
                        request_employee=request_employee,
                        nomenclature=nomenclature,
                        size=item.get('selected_size', ''),
                        quantity=item['selected_quantity']
                    )

        RequestHistory.objects.create(
            request=request_obj,
            status_from='pending',
            status_to='pending',
            changed_by=request.user,
            comment=f'Заявка отредактирована: {title}'
        )

        return Response({'message': 'Request updated successfully', 'request_id': request_obj.request_id})
    @action(detail=True, methods=['post'], permission_classes=[IsEconomicHead])
    def issue_ppe(self, request, pk=None):
        """Выдать СИЗ (статус delivered -> completed)"""
        request_obj = self.get_object()

        if request_obj.status != 'delivered':
            return Response(
                {'error': 'Можно выдавать СИЗ только после доставки'},
                status=400
            )

        issue_date = request.data.get('issue_date', date.today().isoformat())
        comment = request.data.get('comment', '')
        issued_items = request.data.get('issued_items', [])

        if not issued_items:
            return Response({'error': 'Отметьте хотя бы одну позицию для выдачи'}, status=400)

        issued_by = AppUser.objects.get(login=request.user.username)
        issue_date_obj = datetime.strptime(issue_date, '%Y-%m-%d').date()

        for item in issued_items:
            try:
                employee = Employee.objects.get(employee_id=item['employee_id'])
                nomenclature = Nomenclature.objects.get(nomenclature_id=item['nomenclature_id'])
                period_months = item.get('period_months', 12)

                from dateutil.relativedelta import relativedelta
                next_issue_date = issue_date_obj + relativedelta(months=period_months)

                PPEIssueRecord.objects.create(
                    employee=employee,
                    nomenclature=nomenclature,
                    issue_date=issue_date_obj,
                    size=item.get('size', ''),
                    quantity=item['quantity'],
                    next_issue_date=next_issue_date,
                    issued_by=issued_by,
                    comment=f'Выдача по заявке {request_obj.request_number}'
                )
            except Exception as e:
                print(f"Error creating issue: {e}")

        # Меняем статус на completed
        old_status = request_obj.status
        request_obj.status = 'completed'
        request_obj.comment = comment
        request_obj.save()

        RequestHistory.objects.create(
            request=request_obj,
            status_from=old_status,
            status_to='completed',
            changed_by=request.user,
            comment=f'Выданы СИЗ: {comment}'
        )

        return Response(RequestSerializer(request_obj).data)

    @action(detail=True, methods=['post'], permission_classes=[IsEconomicHead])
    def mark_ordered(self, request, pk=None):
        """Отметить что заказ оформлен (статус approved -> ordered)"""
        request_obj = self.get_object()

        # Проверяем статус
        if request_obj.status != 'approved':
            return Response(
                {'error': 'Можно отметить заказ только для одобренных заявок'},
                status=400
            )

        # Получаем данные из запроса
        supplier_name = request.data.get('supplier_name')
        order_price = request.data.get('order_price')
        comment = request.data.get('comment', '')
        ordered_items = request.data.get('ordered_items', [])

        # Проверяем обязательные поля
        if not supplier_name:
            return Response({'error': 'Укажите поставщика'}, status=400)

        if not ordered_items:
            return Response({'error': 'Отметьте хотя бы одну позицию как заказанную'}, status=400)

        # Обновляем заявку
        old_status = request_obj.status
        request_obj.status = 'ordered'
        request_obj.supplier_name = supplier_name
        if order_price:
            request_obj.order_price = order_price
        request_obj.order_date = timezone.now()
        if comment:
            request_obj.comment = comment
        request_obj.save()

        # Создаем историю
        RequestHistory.objects.create(
            request=request_obj,
            status_from=old_status,
            status_to='ordered',
            changed_by=request.user,
            comment=f'Оформлен заказ у поставщика {supplier_name}' + (
                f' на сумму {order_price} руб.' if order_price else '')
        )

        # Здесь можно сохранить информацию о том, какие именно позиции были заказаны
        # Создайте отдельную таблицу RequestOrderedItems если нужно

        return Response({
            'status': 'ordered',
            'message': 'Заказ успешно оформлен'
        })
    @action(detail=True, methods=['post'], permission_classes=[IsEconomicHead])
    def mark_delivered(self, request, pk=None):
        """Отметить доставку (статус ordered -> delivered)"""
        request_obj = self.get_object()

        if request_obj.status != 'ordered':
            return Response(
                {'error': 'Можно отметить доставку только для заказанных заявок'},
                status=400
            )

        comment = request.data.get('comment', '')
        delivered_items = request.data.get('delivered_items', [])

        if not delivered_items:
            return Response({'error': 'Отметьте хотя бы одну позицию как доставленную'}, status=400)

        old_status = request_obj.status
        request_obj.status = 'delivered'
        request_obj.comment = comment
        request_obj.save()

        RequestHistory.objects.create(
            request=request_obj,
            status_from=old_status,
            status_to='delivered',
            changed_by=request.user,
            comment=f'Отмечена доставка: {comment}'
        )

        return Response(RequestSerializer(request_obj).data)

    @action(detail=True, methods=['post'], permission_classes=[IsEconomicHead])
    def issue_ppe(self, request, pk=None):
        """Выдать СИЗ (статус delivered -> completed)"""
        request_obj = self.get_object()

        if request_obj.status != 'delivered':
            return Response(
                {'error': 'Можно выдавать СИЗ только после доставки'},
                status=400
            )

        issue_date = request.data.get('issue_date', date.today().isoformat())
        comment = request.data.get('comment', '')
        issued_items = request.data.get('issued_items', [])

        if not issued_items:
            return Response({'error': 'Отметьте хотя бы одну позицию для выдачи'}, status=400)

        issued_by = AppUser.objects.get(login=request.user.username)
        issue_date_obj = datetime.strptime(issue_date, '%Y-%m-%d').date()

        # Создаем записи о выдаче СИЗ
        from dateutil.relativedelta import relativedelta

        for item in issued_items:
            try:
                employee = Employee.objects.get(employee_id=item['employee_id'])
                nomenclature = Nomenclature.objects.get(nomenclature_id=item['nomenclature_id'])
                period_months = item.get('period_months', 12)

                next_issue_date = issue_date_obj + relativedelta(months=period_months)

                PPEIssueRecord.objects.create(
                    employee=employee,
                    nomenclature=nomenclature,
                    issue_date=issue_date_obj,
                    size=item.get('size', ''),
                    quantity=item['quantity'],
                    next_issue_date=next_issue_date,
                    issued_by=issued_by,
                    comment=f'Выдача по заявке {request_obj.request_number}'
                )
            except Exception as e:
                print(f"Error creating issue: {e}")

        old_status = request_obj.status
        request_obj.status = 'completed'
        request_obj.comment = comment
        request_obj.save()

        RequestHistory.objects.create(
            request=request_obj,
            status_from=old_status,
            status_to='completed',
            changed_by=request.user,
            comment=f'Выданы СИЗ: {comment}'
        )

        return Response(RequestSerializer(request_obj).data)
    @action(detail=True, methods=['get'])
    def request_details(self, request, pk=None):
        """Получить детальную информацию о заявке (сотрудники и СИЗ)"""
        try:
            # Используем get_object_or_404 для лучшей обработки
            from django.shortcuts import get_object_or_404
            request_obj = get_object_or_404(Request, pk=pk)

            # Базовая информация о заявке
            result = {
                'request_id': request_obj.request_id,
                'request_number': request_obj.request_number,
                'title': request_obj.title or '',
                'description': request_obj.description or '',
                'status': request_obj.status,
                'status_display': request_obj.get_status_display(),
                'created_at': request_obj.created_at,
                'requester_name': '',
                'shop_name': '',
                'comment': request_obj.comment or '',
                'employees': []
            }

            # Получаем информацию о заявителе
            if request_obj.user:
                if request_obj.user.employee:
                    result['requester_name'] = request_obj.user.employee.full_name
                else:
                    result['requester_name'] = request_obj.user.login

            # Получаем информацию о цехе
            if request_obj.shop:
                result['shop_name'] = request_obj.shop.title

            # Получаем сотрудников и их СИЗ
            request_employees = RequestEmployee.objects.filter(request=request_obj).select_related('employee',
                                                                                                   'employee__position')

            for req_emp in request_employees:
                if req_emp.employee:
                    employee_data = {
                        'employee_id': req_emp.employee.employee_id,
                        'full_name': req_emp.employee.full_name,
                        'position_name': req_emp.employee.position.title if req_emp.employee.position else '',
                        'height': req_emp.height,
                        'items': []
                    }

                    # Получаем СИЗ для сотрудника
                    items = RequestEmployeeItem.objects.filter(request_employee=req_emp).select_related('nomenclature')
                    for item in items:
                        if item.nomenclature:
                            employee_data['items'].append({
                                'nomenclatureID': item.nomenclature.nomenclature_id,  # ← эта строка должна быть
                                'nomenclature_title': item.nomenclature.title,
                                'size': item.size or '',
                                'quantity': float(item.quantity) if item.quantity else 0,
                                'unit': item.nomenclature.unit or '',
                            })

                    result['employees'].append(employee_data)

            return Response(result)

        except Exception as e:
            print(f"Error in request_details for pk={pk}: {str(e)}")
            import traceback
            traceback.print_exc()
            return Response({'error': str(e)}, status=404)

    @action(detail=True, methods=['post'], permission_classes=[IsSafetyOfficer])
    def approve_by_safety(self, request, pk=None):
        """Одобрить заявку охраной труда"""
        request_obj = self.get_object()

        if request_obj.status != 'pending':
            return Response(
                {'error': 'Можно одобрить только заявки со статусом "На рассмотрении (Охрана труда)"'},
                status=400
            )

        old_status = request_obj.status
        request_obj.status = 'hr_approved'
        request_obj.comment = request.data.get('comment', '')
        request_obj.save()

        RequestHistory.objects.create(
            request=request_obj,
            status_from=old_status,
            status_to='hr_approved',
            changed_by=request.user,
            comment=f'Одобрено охраной труда: {request.data.get("comment", "")}'
        )

        return Response(RequestSerializer(request_obj).data)

    @action(detail=True, methods=['post'], permission_classes=[IsSafetyOfficer])
    def reject_by_safety(self, request, pk=None):
        """Отклонить заявку охраной труда"""
        request_obj = self.get_object()

        if request_obj.status != 'pending':
            return Response(
                {'error': 'Можно отклонить только заявки со статусом "На рассмотрении (Охрана труда)"'},
                status=400
            )

        comment = request.data.get('comment', '')
        request_obj.status = 'rejected'
        request_obj.comment = comment
        request_obj.save()

        RequestHistory.objects.create(
            request=request_obj,
            status_from='pending',
            status_to='rejected',
            changed_by=request.user,
            comment=f'Отклонено охраной труда: {comment}'
        )

        return Response(RequestSerializer(request_obj).data)

    @action(detail=False, methods=['get'], permission_classes=[IsEconomicHead])
    def economic_pending_requests(self, request):
        """Получить заявки, одобренные охраной труда, для хоз. отдела"""
        requests = Request.objects.filter(status='hr_approved').order_by('-created_at')
        print(f"Found {requests.count()} requests with status 'hr_approved'")  # Добавьте для отладки
        serializer = RequestSerializer(requests, many=True)
        return Response(serializer.data)

    # Добавьте временный print в метод safety_pending_requests
    @action(detail=False, methods=['get'], permission_classes=[IsSafetyOfficer])
    def safety_pending_requests(self, request):
        """Получить заявки на рассмотрении для охраны труда"""
        requests = Request.objects.filter(status='pending').order_by('-created_at')
        print(f"Found {requests.count()} pending requests for safety officer")
        for req in requests:
            print(f"  - {req.request_number}: {req.status}")
        serializer = RequestSerializer(requests, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'], permission_classes=[IsSafetyOfficer])
    def reject_by_safety(self, request, pk=None):
        """Отклонить заявку охраной труда"""
        request_obj = self.get_object()

        if request_obj.status != 'pending':
            return Response(
                {'error': 'Можно отклонить только заявки со статусом "На рассмотрении (Охрана труда)"'},
                status=400
            )

        comment = request.data.get('comment', '')
        old_status = request_obj.status
        request_obj.status = 'rejected'
        request_obj.comment = comment
        request_obj.save()

        RequestHistory.objects.create(
            request=request_obj,
            status_from=old_status,
            status_to='rejected',
            changed_by=request.user,
            comment=f'Отклонено охраной труда: {comment}'
        )

        return Response(RequestSerializer(request_obj).data)

    @action(detail=True, methods=['post'], permission_classes=[IsEconomicHead])
    def approve_by_economic(self, request, pk=None):
        """Одобрить заявку хоз. отделом"""
        request_obj = self.get_object()

        # Этот метод должен принимать только заявки со статусом 'hr_approved'
        if request_obj.status != 'hr_approved':
            return Response(
                {'error': 'Можно одобрить только заявки со статусом "Одобрено охраной труда"'},
                status=400
            )

        old_status = request_obj.status
        request_obj.status = 'approved'
        request_obj.comment = request.data.get('comment', '')
        request_obj.save()

        print(f"Request {request_obj.request_number}: {old_status} -> {request_obj.status}")  # Отладка

        RequestHistory.objects.create(
            request=request_obj,
            status_from=old_status,
            status_to='approved',
            changed_by=request.user,
            comment=f'Одобрено хоз. отделом: {request.data.get("comment", "")}'
        )

        return Response(RequestSerializer(request_obj).data)

    @action(detail=True, methods=['post'], permission_classes=[IsEconomicHead])
    def make_order(self, request, pk=None):
        """Оформить заказ (статус approved -> ordered)"""
        request_obj = self.get_object()

        if request_obj.status != 'approved':
            return Response(
                {'error': 'Можно оформить заказ только для одобренных заявок'},
                status=400
            )

        supplier_name = request.data.get('supplier_name')
        order_price = request.data.get('order_price')
        comment = request.data.get('comment', '')

        if not supplier_name or not order_price:
            return Response({'error': 'Укажите поставщика и стоимость'}, status=400)

        old_status = request_obj.status
        request_obj.status = 'ordered'
        request_obj.supplier_name = supplier_name
        request_obj.order_price = order_price
        request_obj.order_date = timezone.now()
        request_obj.comment = comment
        request_obj.save()

        RequestHistory.objects.create(
            request=request_obj,
            status_from=old_status,
            status_to='ordered',
            changed_by=request.user,
            comment=f'Оформлен заказ у поставщика {supplier_name} на сумму {order_price} руб.'
        )

        return Response(RequestSerializer(request_obj).data)

    @action(detail=True, methods=['post'], permission_classes=[IsEconomicHead])
    def mark_delivered(self, request, pk=None):
        """Отметить доставку (статус ordered -> delivered)"""
        request_obj = self.get_object()

        if request_obj.status != 'ordered':
            return Response(
                {'error': 'Можно отметить доставку только для заказанных заявок'},
                status=400
            )

        comment = request.data.get('comment', '')

        old_status = request_obj.status
        request_obj.status = 'delivered'
        request_obj.comment = comment
        request_obj.save()

        RequestHistory.objects.create(
            request=request_obj,
            status_from=old_status,
            status_to='delivered',
            changed_by=request.user,
            comment=f'Отмечена доставка: {comment}'
        )

        return Response(RequestSerializer(request_obj).data)

    @action(detail=True, methods=['post'], permission_classes=[IsEconomicHead])
    def reject_by_economic(self, request, pk=None):
        """Отклонить заявку хоз. отделом"""
        request_obj = self.get_object()

        if request_obj.status != 'hr_approved':
            return Response(
                {'error': 'Можно отклонить только заявки со статусом "Одобрено охраной труда"'},
                status=400
            )

        comment = request.data.get('comment', '')
        old_status = request_obj.status
        request_obj.status = 'rejected'
        request_obj.comment = comment
        request_obj.save()

        RequestHistory.objects.create(
            request=request_obj,
            status_from=old_status,
            status_to='rejected',
            changed_by=request.user,
            comment=f'Отклонено хоз. отделом: {comment}'
        )

        return Response(RequestSerializer(request_obj).data)

    @action(detail=True, methods=['post'], permission_classes=[IsEconomicHead])
    def make_order(self, request, pk=None):
        """Сделать заказ (после одобрения хоз. отделом)"""
        request_obj = self.get_object()

        if request_obj.status != 'approved':
            return Response(
                {'error': 'Заказ можно сделать только для полностью одобренных заявок'},
                status=400
            )

        serializer = RequestOrderSerializer(data=request.data)
        if serializer.is_valid():
            old_status = request_obj.status
            request_obj.status = 'ordered'
            request_obj.supplier_name = serializer.validated_data['supplier_name']
            request_obj.order_price = serializer.validated_data['order_price']
            request_obj.order_date = timezone.now()
            request_obj.comment = serializer.validated_data.get('comment', '')
            request_obj.save()

            RequestHistory.objects.create(
                request=request_obj,
                status_from=old_status,
                status_to='ordered',
                changed_by=request.user,
                comment=f'Заказ у поставщика {request_obj.supplier_name} на сумму {request_obj.order_price} руб.'
            )
            return Response(RequestSerializer(request_obj).data)
        return Response(serializer.errors, status=400)

    @action(detail=True, methods=['post'], permission_classes=[IsEconomicHead])
    def complete(self, request, pk=None):
        """Завершить заявку (статус -> completed)"""
        request_obj = self.get_object()

        if request_obj.status != 'ordered':
            return Response(
                {'error': 'Можно завершить только заявки со статусом "Заказ сделан"'},
                status=400
            )

        old_status = request_obj.status
        request_obj.status = 'completed'
        request_obj.comment = request.data.get('comment', request_obj.comment or '')
        request_obj.save()

        RequestHistory.objects.create(
            request=request_obj,
            status_from=old_status,
            status_to='completed',
            changed_by=request.user,
            comment=request.data.get('comment', 'Заявка выполнена')
        )

        return Response({'status': 'completed', 'message': 'Заявка завершена'})
    # Обновите также метод get_queryset для фильтрации по ролям
    def get_queryset(self):
        user = self.request.user
        print(f"User: {user.username}, is_authenticated: {user.is_authenticated}")

        try:
            app_user = AppUser.objects.get(login=user.username)
            print(f"AppUser role: {app_user.role}")
        except AppUser.DoesNotExist:
            print(f"AppUser not found for {user.username}")
            return Request.objects.none()

        # Администратор видит все заявки
        if user.is_superuser:
            print("Superuser - returning all requests")
            return Request.objects.all().order_by('-created_at')

        # Охрана труда видит заявки в статусе 'pending'
        if app_user.role == 'safety_officer':
            print("Safety officer - returning pending requests")
            return Request.objects.filter(status='pending').order_by('-created_at')

        # Хоз. отдел видит заявки со статусами 'hr_approved', 'approved', 'ordered', 'delivered', 'completed'
        if app_user.role == 'economic_head':
            print("Economic head - returning requests for economic")
            return Request.objects.filter(
                status__in=['hr_approved', 'approved', 'ordered', 'delivered', 'completed']
            ).order_by('-created_at')

        # Остальные видят только свои заявки
        print(f"Regular user - returning own requests")
        return Request.objects.filter(user=app_user).order_by('-created_at')
    @action(detail=True, methods=['post'], permission_classes=[IsAdminOrEconomicHead])
    def approve(self, request, pk=None):
        """Одобрить заявку (админ или начальник хоз. отдела)"""
        request_obj = self.get_object()
        if request_obj.status != 'pending':
            return Response(
                {'error': 'Можно одобрить только заявки со статусом "На рассмотрении"'},
                status=400
            )

        serializer = RequestApproveSerializer(data=request.data)
        if serializer.is_valid():
            old_status = request_obj.status
            request_obj.status = 'approved'
            request_obj.comment = serializer.validated_data.get('comment', '')
            request_obj.save()

            RequestHistory.objects.create(
                request=request_obj,
                status_from=old_status,
                status_to='approved',
                changed_by=request.user,
                comment=serializer.validated_data.get('comment', '')
            )
            return Response(RequestSerializer(request_obj).data)
        return Response(serializer.errors, status=400)

    @action(detail=True, methods=['post'], permission_classes=[IsAdminOrEconomicHead])
    def reject(self, request, pk=None):
        """Отклонить заявку (админ или начальник хоз. отдела)"""
        request_obj = self.get_object()
        if request_obj.status != 'pending':
            return Response(
                {'error': 'Можно отклонить только заявки со статусом "На рассмотрении"'},
                status=400
            )

        comment = request.data.get('comment', '')
        old_status = request_obj.status
        request_obj.status = 'rejected'
        request_obj.comment = comment
        request_obj.save()

        RequestHistory.objects.create(
            request=request_obj,
            status_from=old_status,
            status_to='rejected',
            changed_by=request.user,
            comment=comment
        )
        return Response(RequestSerializer(request_obj).data)

    @action(detail=True, methods=['post'], permission_classes=[IsAdminOrEconomicHead])
    def make_order(self, request, pk=None):
        """Сделать заказ (админ или начальник хоз. отдела)"""
        request_obj = self.get_object()
        if request_obj.status != 'approved':
            return Response(
                {'error': 'Заказ можно сделать только для одобренных заявок'},
                status=400
            )

        serializer = RequestOrderSerializer(data=request.data)
        if serializer.is_valid():
            old_status = request_obj.status
            request_obj.status = 'ordered'
            request_obj.supplier_name = serializer.validated_data['supplier_name']
            request_obj.order_price = serializer.validated_data['order_price']
            request_obj.order_date = timezone.now()
            request_obj.comment = serializer.validated_data.get('comment', '')
            request_obj.save()

            RequestHistory.objects.create(
                request=request_obj,
                status_from=old_status,
                status_to='ordered',
                changed_by=request.user,
                comment=f'Заказ у поставщика {request_obj.supplier_name} на сумму {request_obj.order_price} руб.'
            )
            return Response(RequestSerializer(request_obj).data)
        return Response(serializer.errors, status=400)

    @action(detail=True, methods=['post'], permission_classes=[IsAdminOrEconomicHead])
    def complete(self, request, pk=None):
        """Завершить заявку (админ или начальник хоз. отдела)"""
        request_obj = self.get_object()
        if request_obj.status != 'ordered':
            return Response(
                {'error': 'Можно завершить только заявки со статусом "Заказ сделан"'},
                status=400
            )

        old_status = request_obj.status
        request_obj.status = 'completed'
        request_obj.save()

        RequestHistory.objects.create(
            request=request_obj,
            status_from=old_status,
            status_to='completed',
            changed_by=request.user,
            comment='Заявка выполнена'
        )
        return Response(RequestSerializer(request_obj).data)

    @action(detail=True, methods=['put'], permission_classes=[IsDepartmentHead])
    def update_request(self, request, pk=None):
        """Редактировать заявку (только создатель, только в статусе pending)"""
        request_obj = self.get_object()

        if request_obj.status != 'pending':
            return Response(
                {'error': 'Можно редактировать только заявки в статусе "На рассмотрении"'},
                status=400
            )

        # Проверяем, что пользователь - создатель заявки
        try:
            app_user = AppUser.objects.get(login=request.user.username)
            if request_obj.user != app_user and not request.user.is_superuser:
                return Response(
                    {'error': 'Вы можете редактировать только свои заявки'},
                    status=403
                )
        except AppUser.DoesNotExist:
            return Response({'error': 'User not found'}, status=404)

        serializer = RequestCreateSerializer(data=request.data)
        if serializer.is_valid():
            request_obj.title = serializer.validated_data.get('title', request_obj.title)
            request_obj.description = serializer.validated_data.get('description', request_obj.description)
            request_obj.quantity = serializer.validated_data.get('quantity', request_obj.quantity)
            request_obj.unit = serializer.validated_data.get('unit', request_obj.unit)
            request_obj.save()

            RequestHistory.objects.create(
                request=request_obj,
                status_from=request_obj.status,
                status_to='pending',
                changed_by=request.user,
                comment='Заявка отредактирована'
            )
            return Response(RequestSerializer(request_obj).data)
        return Response(serializer.errors, status=400)

    @action(detail=True, methods=['post'], permission_classes=[IsDepartmentHead])
    def cancel_request(self, request, pk=None):
        """Отозвать заявку (только создатель, только в статусе pending)"""
        request_obj = self.get_object()

        if request_obj.status != 'pending':
            return Response(
                {'error': 'Можно отозвать только заявки в статусе "На рассмотрении"'},
                status=400
            )

        # Проверяем, что пользователь - создатель заявки
        try:
            app_user = AppUser.objects.get(login=request.user.username)
            if request_obj.user != app_user and not request.user.is_superuser:
                return Response(
                    {'error': 'Вы можете отозвать только свои заявки'},
                    status=403
                )
        except AppUser.DoesNotExist:
            return Response({'error': 'User not found'}, status=404)

        old_status = request_obj.status
        request_obj.status = 'cancelled'
        request_obj.comment = request.data.get('comment', 'Заявка отозвана')
        request_obj.save()

        RequestHistory.objects.create(
            request=request_obj,
            status_from=old_status,
            status_to='cancelled',
            changed_by=request.user,
            comment=request.data.get('comment', 'Заявка отозвана')
        )
        return Response(RequestSerializer(request_obj).data)

    @action(detail=False, methods=['get'])
    def my_requests(self, request):
        """Получить мои заявки"""
        try:
            app_user = AppUser.objects.get(login=request.user.username)
            requests = Request.objects.filter(user=app_user).order_by('-created_at')
            serializer = RequestSerializer(requests, many=True)
            return Response(serializer.data)
        except AppUser.DoesNotExist:
            return Response([])

    @action(detail=False, methods=['get'], permission_classes=[IsAdminOrEconomicHead])
    def pending_requests(self, request):
        """Получить заявки на рассмотрении (админ или начальник хоз. отдела)"""
        requests = Request.objects.filter(status='pending').order_by('-created_at')
        serializer = RequestSerializer(requests, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], permission_classes=[IsAdminOrEconomicHead])
    def all_requests_for_economic(self, request):
        """Получить все заявки (админ или начальник хоз. отдела)"""
        requests = Request.objects.all().order_by('-created_at')
        serializer = RequestSerializer(requests, many=True)
        return Response(serializer.data)

    def get_serializer_class(self):
        if self.action == 'create':
            return RequestCreateSerializer
        return RequestSerializer

    def perform_create(self, serializer):
        app_user = AppUser.objects.get(login=self.request.user.username)
        year = datetime.now().year
        last_request = Request.objects.order_by('-request_id').first()
        if last_request and last_request.request_number:
            try:
                last_num = int(last_request.request_number.split('-')[-1])
                new_num = last_num + 1
            except (ValueError, IndexError):
                new_num = 1
        else:
            new_num = 1
        request_number = f"REQ-{year}-{new_num:06d}"

        request_obj = serializer.save(
            user=app_user,
            shop=app_user.shop,
            request_number=request_number,
            status='pending'
        )
        RequestHistory.objects.create(
            request=request_obj,
            status_from=None,
            status_to='pending',
            changed_by=self.request.user,
            comment='Заявка создана'
        )

    @action(detail=True, methods=['put'], permission_classes=[IsAuthenticated])
    def update_request(self, request, pk=None):
        """Редактировать заявку (только для заявок в статусе pending)"""
        request_obj = self.get_object()

        # Проверяем, что заявка в статусе "На рассмотрении"
        if request_obj.status != 'pending':
            return Response(
                {'error': 'Можно редактировать только заявки в статусе "На рассмотрении"'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Проверяем, что пользователь - создатель заявки
        try:
            app_user = AppUser.objects.get(login=request.user.username)
            if request_obj.user != app_user:
                return Response(
                    {'error': 'Вы можете редактировать только свои заявки'},
                    status=status.HTTP_403_FORBIDDEN
                )
        except AppUser.DoesNotExist:
            return Response({'error': 'User not found'}, status=404)

        serializer = RequestCreateSerializer(data=request.data)
        if serializer.is_valid():
            request_obj.title = serializer.validated_data.get('title', request_obj.title)
            request_obj.description = serializer.validated_data.get('description', request_obj.description)
            request_obj.quantity = serializer.validated_data.get('quantity', request_obj.quantity)
            request_obj.unit = serializer.validated_data.get('unit', request_obj.unit)
            request_obj.save()

            RequestHistory.objects.create(
                request=request_obj,
                status_from=request_obj.status,
                status_to='pending',
                changed_by=request.user,
                comment='Заявка отредактирована'
            )

            return Response(RequestSerializer(request_obj).data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'], permission_classes=[IsAuthenticated])
    def cancel_request(self, request, pk=None):
        """Отозвать заявку (только для заявок в статусе pending)"""
        request_obj = self.get_object()

        # Проверяем, что заявка в статусе "На рассмотрении"
        if request_obj.status != 'pending':
            return Response(
                {'error': 'Можно отозвать только заявки в статусе "На рассмотрении"'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Проверяем, что пользователь - создатель заявки
        try:
            app_user = AppUser.objects.get(login=request.user.username)
            if request_obj.user != app_user:
                return Response(
                    {'error': 'Вы можете отозвать только свои заявки'},
                    status=status.HTTP_403_FORBIDDEN
                )
        except AppUser.DoesNotExist:
            return Response({'error': 'User not found'}, status=404)

        old_status = request_obj.status
        request_obj.status = 'cancelled'
        request_obj.comment = request.data.get('comment', 'Заявка отозвана')
        request_obj.save()

        RequestHistory.objects.create(
            request=request_obj,
            status_from=old_status,
            status_to='cancelled',
            changed_by=request.user,
            comment=request.data.get('comment', 'Заявка отозвана')
        )

        return Response(RequestSerializer(request_obj).data)

    @action(detail=True, methods=['post'])
    def cancel_request(self, request, pk=None):
        """Отозвать заявку (только для заявок в статусе pending)"""
        request_obj = self.get_object()

        # Проверяем, что заявка в статусе "На рассмотрении"
        if request_obj.status != 'pending':
            return Response(
                {'error': 'Можно отозвать только заявки в статусе "На рассмотрении"'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Проверяем, что пользователь - создатель заявки
        try:
            app_user = AppUser.objects.get(login=request.user.username)
            if request_obj.user != app_user:
                return Response(
                    {'error': 'Вы можете отозвать только свои заявки'},
                    status=status.HTTP_403_FORBIDDEN
                )
        except AppUser.DoesNotExist:
            return Response({'error': 'User not found'}, status=404)

        old_status = request_obj.status
        request_obj.status = 'cancelled'
        request_obj.comment = request.data.get('comment', 'Заявка отозвана')
        request_obj.save()

        RequestHistory.objects.create(
            request=request_obj,
            status_from=old_status,
            status_to='cancelled',
            changed_by=request.user,
            comment=request.data.get('comment', 'Заявка отозвана')
        )

        return Response(RequestSerializer(request_obj).data)

    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def all_requests_for_admin(self, request):
        """Получить все заявки (для администратора и проверяющего)"""
        user = request.user

        # Администратор или проверяющий могут видеть все заявки
        if user.is_superuser or user.groups.filter(name='economic_head').exists():
            requests = Request.objects.all().order_by('-created_at')
            serializer = RequestSerializer(requests, many=True)
            return Response(serializer.data)

        return Response(
            {'error': 'Доступ запрещен'},
            status=status.HTTP_403_FORBIDDEN
        )
    @action(detail=True, methods=['post'], permission_classes=[IsEconomicHead])
    def approve(self, request, pk=None):
        request_obj = self.get_object()
        if request_obj.status != 'pending':
            return Response({'error': 'Можно одобрить только заявки со статусом "На рассмотрении"'}, status=400)

        serializer = RequestApproveSerializer(data=request.data)
        if serializer.is_valid():
            old_status = request_obj.status
            request_obj.status = 'approved'
            request_obj.comment = serializer.validated_data.get('comment', '')
            request_obj.save()

            RequestHistory.objects.create(
                request=request_obj,
                status_from=old_status,
                status_to='approved',
                changed_by=request.user,
                comment=serializer.validated_data.get('comment', '')
            )
            return Response(RequestSerializer(request_obj).data)
        return Response(serializer.errors, status=400)

    @action(detail=True, methods=['post'], permission_classes=[IsEconomicHead])
    def reject(self, request, pk=None):
        request_obj = self.get_object()
        if request_obj.status != 'pending':
            return Response({'error': 'Можно отклонить только заявки со статусом "На рассмотрении"'}, status=400)

        comment = request.data.get('comment', '')
        old_status = request_obj.status
        request_obj.status = 'rejected'
        request_obj.comment = comment
        request_obj.save()

        RequestHistory.objects.create(
            request=request_obj,
            status_from=old_status,
            status_to='rejected',
            changed_by=request.user,
            comment=comment
        )
        return Response(RequestSerializer(request_obj).data)

    @action(detail=True, methods=['post'], permission_classes=[IsEconomicHead])
    def make_order(self, request, pk=None):
        request_obj = self.get_object()
        if request_obj.status != 'approved':
            return Response({'error': 'Заказ можно сделать только для одобренных заявок'}, status=400)

        serializer = RequestOrderSerializer(data=request.data)
        if serializer.is_valid():
            old_status = request_obj.status
            request_obj.status = 'ordered'
            request_obj.supplier_name = serializer.validated_data['supplier_name']
            request_obj.order_price = serializer.validated_data['order_price']
            request_obj.order_date = timezone.now()
            request_obj.comment = serializer.validated_data.get('comment', '')
            request_obj.save()

            RequestHistory.objects.create(
                request=request_obj,
                status_from=old_status,
                status_to='ordered',
                changed_by=request.user,
                comment=f'Заказ у поставщика {request_obj.supplier_name} на сумму {request_obj.order_price} руб.'
            )
            return Response(RequestSerializer(request_obj).data)
        return Response(serializer.errors, status=400)

    @action(detail=True, methods=['post'], permission_classes=[IsEconomicHead])
    def complete(self, request, pk=None):
        request_obj = self.get_object()
        if request_obj.status != 'ordered':
            return Response({'error': 'Можно завершить только заявки со статусом "Заказ сделан"'}, status=400)

        old_status = request_obj.status
        request_obj.status = 'completed'
        request_obj.save()

        RequestHistory.objects.create(
            request=request_obj,
            status_from=old_status,
            status_to='completed',
            changed_by=request.user,
            comment='Заявка выполнена'
        )
        return Response(RequestSerializer(request_obj).data)

    @action(detail=True, methods=['get'])
    def history(self, request, pk=None):
        request_obj = self.get_object()
        history = request_obj.history.all()
        serializer = RequestHistorySerializer(history, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def my_requests(self, request):
        try:
            app_user = AppUser.objects.get(login=request.user.username)
            requests = Request.objects.filter(user=app_user).order_by('-created_at')
            serializer = RequestSerializer(requests, many=True)
            return Response(serializer.data)
        except AppUser.DoesNotExist:
            return Response([])

    @action(detail=False, methods=['get'])
    def pending_requests(self, request):
        """Получить заявки на рассмотрении (для начальника хоз. отдела и админа)"""
        user = request.user

        # Разрешаем доступ для начальника хоз. отдела и администратора
        if not (user.groups.filter(name='economic_head').exists() or user.is_superuser):
            return Response(
                {'error': 'Доступно только для начальника хозяйственного отдела и администратора'},
                status=status.HTTP_403_FORBIDDEN
            )

        # Показываем все заявки со статусом 'pending'
        requests = Request.objects.filter(status='pending').order_by('-created_at')
        serializer = RequestSerializer(requests, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def all_requests_for_economic(self, request):
        """Получить все заявки для хоз. отдела"""
        user = request.user

        # Получаем AppUser для проверки роли
        try:
            app_user = AppUser.objects.get(login=user.username)
            user_role = app_user.role
        except AppUser.DoesNotExist:
            return Response({'error': 'User not found'}, status=404)

        # Разрешаем доступ для начальника хоз. отдела и администратора
        if user_role not in ['economic_head', 'admin'] and not user.is_superuser:
            return Response(
                {'error': 'Доступ запрещен. Требуется роль "economic_head" или "admin"'},
                status=403
            )

        # Показываем заявки со статусами, которые актуальны для хоз. отдела
        requests = Request.objects.filter(
            status__in=['hr_approved', 'approved', 'ordered', 'delivered', 'completed']
        ).order_by('-created_at')

        serializer = RequestSerializer(requests, many=True)
        return Response(serializer.data)
# ==================== USER VIEWS ====================

class UserViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=['get'])
    def me(self, request):
        serializer = self.get_serializer(request.user)
        try:
            app_user = AppUser.objects.get(login=request.user.username)
            data = serializer.data
            data['role'] = app_user.role  # Убедитесь, что это поле существует
            data['shop_id'] = app_user.shop.shop_id if app_user.shop else None
            data['shop_name'] = app_user.shop.title if app_user.shop else ''
            data[
                'full_name'] = app_user.employee.full_name if app_user.employee else f"{request.user.first_name} {request.user.last_name}"
            print(f"User role for {request.user.username}: {app_user.role}")  # Добавьте для отладки
            return Response(data)
        except AppUser.DoesNotExist:
            return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def by_role(self, request):
        role = request.query_params.get('role', None)
        if role:
            app_users = AppUser.objects.filter(role=role)
            users = User.objects.filter(username__in=[au.login for au in app_users])
            serializer = self.get_serializer(users, many=True)
            return Response(serializer.data)
        return Response([])


@api_view(['POST'])
@permission_classes([AllowAny])
def simple_auth(request):
    username = request.data.get('username')
    password = request.data.get('password')

    print(f"=== AUTH ATTEMPT ===")
    print(f"Username: {username}")

    try:
        app_user = AppUser.objects.filter(login=username).first()
        if not app_user:
            print(f"User not found: {username}")
            return Response({'error': 'User not found'}, status=401)

        print(f"User found: {username}, role: {app_user.role}")

        if check_password(password, app_user.password):
            print("Password correct!")

            django_user, created = User.objects.get_or_create(
                username=app_user.login,
                defaults={'email': f"{app_user.login}@example.com", 'is_active': True}
            )

            if app_user.employee:
                django_user.first_name = app_user.employee.first_name
                django_user.last_name = app_user.employee.last_name
                django_user.save()

            token, _ = Token.objects.get_or_create(user=django_user)

            print(f"Token created: {token.key[:20]}...")

            return Response({
                'token': token.key,
                'username': django_user.username,
                'is_superuser': django_user.is_superuser,
                'user_id': django_user.id,
                'role': app_user.role
            })
        else:
            print("Password incorrect!")
            return Response({'error': 'Invalid password'}, status=401)
    except Exception as e:
        print(f"Auth error: {str(e)}")
        import traceback
        traceback.print_exc()
        return Response({'error': str(e)}, status=500)

# ==================== PPE VIEWS ====================
class PPEViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=['get'])
    def get_employee_ppe(self, request):
        """Получить список СИЗ для сотрудника по нормам выдачи с информацией о сроках"""
        employee_id = request.query_params.get('employee_id')
        if not employee_id:
            return Response({'error': 'employee_id required'}, status=400)

        try:
            employee = Employee.objects.get(employee_id=employee_id)
        except Employee.DoesNotExist:
            return Response({'error': 'Employee not found'}, status=404)

        # Используем новую модель PPEIssueStandard
        standards = PPEIssueStandard.objects.filter(
            position=employee.position,
            is_active=True
        ).select_related('nomenclature')

        result = []
        today = date.today()

        for standard in standards:
            # Проверяем последнюю выдачу
            last_issue = PPEIssueRecord.objects.filter(
                employee=employee,
                nomenclature=standard.nomenclature
            ).order_by('-issue_date').first()

            can_order = True
            last_issue_date = None
            days_until_expiry = None
            status_color = '#4caf50'  # зеленый по умолчанию
            status_text = 'Срок не истек'

            if last_issue:
                last_issue_date = last_issue.issue_date
                # Рассчитываем количество дней до следующей выдачи
                if last_issue.next_issue_date:
                    days_until_expiry = (last_issue.next_issue_date - today).days

                    if days_until_expiry <= 0:
                        can_order = True
                        status_color = '#f44336'  # красный - просрочено
                        status_text = 'СРОЧНО! Требуется замена'
                    elif days_until_expiry <= 60:  # менее 2 месяцев
                        can_order = True
                        status_color = '#f44336'  # красный
                        status_text = f'Срочно! Осталось {days_until_expiry} дн.'
                    elif days_until_expiry <= 150:  # до 5 месяцев
                        can_order = True
                        status_color = '#ff9800'  # желтый
                        status_text = f'Скоро замена, осталось {days_until_expiry} дн.'
                    else:
                        can_order = True
                        status_color = '#4caf50'  # зеленый
                        status_text = f'Срок не истек, осталось {days_until_expiry} дн.'
                else:
                    # Если нет даты следующей выдачи, считаем от даты выдачи + период
                    from dateutil.relativedelta import relativedelta
                    expiry_date = last_issue_date + relativedelta(months=standard.period_months)
                    days_until_expiry = (expiry_date - today).days

                    if days_until_expiry <= 0:
                        can_order = True
                        status_color = '#f44336'
                        status_text = 'СРОЧНО! Требуется замена'
                    elif days_until_expiry <= 60:
                        can_order = True
                        status_color = '#f44336'
                        status_text = f'Срочно! Осталось {days_until_expiry} дн.'
                    elif days_until_expiry <= 150:
                        can_order = True
                        status_color = '#ff9800'
                        status_text = f'Скоро замена, осталось {days_until_expiry} дн.'
                    else:
                        can_order = True
                        status_color = '#4caf50'
                        status_text = f'Срок не истек, осталось {days_until_expiry} дн.'

            result.append({
                'nomenclature_id': standard.nomenclature.nomenclature_id,
                'nomenclature_title': standard.nomenclature.title,
                'unit': standard.nomenclature.unit,
                'standard_quantity': float(standard.quantity),
                'period_months': standard.period_months,
                'last_issue_date': last_issue_date,
                'can_order': can_order,
                'selected_size': None,
                'selected_quantity': float(standard.quantity) if can_order else 0,
                'days_until_expiry': days_until_expiry,
                'status_color': status_color,
                'status_text': status_text,
            })

        return Response(result)
    @action(detail=False, methods=['get'])
    def get_available_employees(self, request):
        """Получить список сотрудников"""
        try:
            app_user = AppUser.objects.get(login=request.user.username)

            # Если пользователь администратор или имеет специальные права - показываем всех сотрудников
            if request.user.is_superuser or app_user.role in ['admin', 'economic_head', 'safety_officer']:
                print(f"User {app_user.login} with role {app_user.role} - showing ALL employees")
                employees = Employee.objects.filter(is_active=True).select_related('position', 'shop')
            else:
                print(f"User {app_user.login} with role {app_user.role} - showing only shop employees")
                employees = Employee.objects.filter(
                    shop=app_user.shop,
                    is_active=True
                ).select_related('position')

            result = []
            for emp in employees:
                result.append({
                    'employee_id': emp.employee_id,
                    'full_name': emp.full_name,
                    'position_name': emp.position.title,
                    'gender': emp.gender,
                    'heightcm': emp.heightcm,
                    'clothing_size': emp.clothing_size,
                    'shoesize': emp.shoesize,
                    'headsize': emp.headsize,
                    'shop_name': emp.shop.title if emp.shop else None,
                })
            print(f"Returning {len(result)} employees")
            return Response(result)
        except AppUser.DoesNotExist:
            print(f"User not found: {request.user.username}")
            return Response([])
# ==================== FULL REQUEST VIEWS ====================

class FullRequestViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def create(self, request):
        try:
            app_user = AppUser.objects.get(login=request.user.username)
        except AppUser.DoesNotExist:
            return Response({'error': 'User not found'}, status=404)

        data = request.data
        title = data.get('title')
        employees_data = data.get('employees', [])

        if not title:
            return Response({'error': 'Title required'}, status=400)
        if not employees_data:
            return Response({'error': 'At least one employee required'}, status=400)

        year = datetime.now().year
        last_request = Request.objects.order_by('-request_id').first()
        if last_request and last_request.request_number:
            try:
                last_num = int(last_request.request_number.split('-')[-1])
                new_num = last_num + 1
            except (ValueError, IndexError):
                new_num = 1
        else:
            new_num = 1
        request_number = f"REQ-{year}-{new_num:06d}"

        # СОЗДАЕМ ЗАЯВКУ СО СТАТУСОМ 'pending'
        request_obj = Request.objects.create(
            request_number=request_number,
            shop=app_user.shop,
            user=app_user,
            title=title,
            description=data.get('description', ''),
            status='pending',  # Важно! Статус ожидания одобрения охраной труда
            created_at=timezone.now(),
            updated_at=timezone.now()
        )

        RequestHistory.objects.create(
            request=request_obj,
            status_from=None,
            status_to='pending',
            changed_by=request.user,
            comment='Заявка создана'
        )

        total_items = 0
        for emp_data in employees_data:
            employee = Employee.objects.get(employee_id=emp_data['employee_id'])
            request_employee = RequestEmployee.objects.create(
                request=request_obj,
                employee=employee,
                height=emp_data.get('height')
            )
            for item in emp_data['items']:
                if item.get('selected_quantity', 0) > 0:
                    nomenclature = Nomenclature.objects.get(nomenclature_id=item['nomenclature_id'])
                    RequestEmployeeItem.objects.create(
                        request_employee=request_employee,
                        nomenclature=nomenclature,
                        size=item.get('selected_size', ''),
                        quantity=item['selected_quantity']
                    )
                    total_items += 1

        send_websocket_notification(request_obj, 'created')
        excel_file = self.generate_excel_report(request_obj)

        return Response({
            'request_id': request_obj.request_id,
            'request_number': request_obj.request_number,
            'status': request_obj.status,
            'total_employees': len(employees_data),
            'total_items': total_items,
            'excel_report': excel_file
        }, status=201)


    def generate_excel_report(self, request_obj):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Детальный отчет"

        headers = ['№', 'ФИО сотрудника', 'Должность', 'Рост (см)', 'СИЗ', 'Размер', 'Количество', 'Ед. изм.']
        for col, header in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        row = 2
        request_employees = RequestEmployee.objects.filter(request=request_obj)
        for req_emp in request_employees:
            items = RequestEmployeeItem.objects.filter(request_employee=req_emp)
            for item in items:
                ws1.cell(row=row, column=1, value=row - 1)
                ws1.cell(row=row, column=2, value=req_emp.employee.full_name)
                ws1.cell(row=row, column=3, value=req_emp.employee.position.title)
                ws1.cell(row=row, column=4, value=req_emp.height or '')
                ws1.cell(row=row, column=5, value=item.nomenclature.title)
                ws1.cell(row=row, column=6, value=item.size)
                ws1.cell(row=row, column=7, value=float(item.quantity))
                ws1.cell(row=row, column=8, value=item.nomenclature.unit)
                row += 1

        ws2 = wb.create_sheet("Сводный отчет")
        summary_headers = ['СИЗ', 'Размер', 'Общее количество', 'Ед. изм.']
        for col, header in enumerate(summary_headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        summary = {}
        for req_emp in request_employees:
            items = RequestEmployeeItem.objects.filter(request_employee=req_emp)
            for item in items:
                key = (item.nomenclature.title, item.size)
                if key not in summary:
                    summary[key] = {'quantity': 0, 'unit': item.nomenclature.unit}
                summary[key]['quantity'] += float(item.quantity)

        row = 2
        for (name, size), data in summary.items():
            ws2.cell(row=row, column=1, value=name)
            ws2.cell(row=row, column=2, value=size)
            ws2.cell(row=row, column=3, value=data['quantity'])
            ws2.cell(row=row, column=4, value=data['unit'])
            row += 1

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        import base64
        return base64.b64encode(output.getvalue()).decode('utf-8')

    @action(detail=False, methods=['post'])
    def approve_request(self, request):
        request_id = request.data.get('request_id')
        try:
            request_obj = Request.objects.get(request_id=request_id)
        except Request.DoesNotExist:
            return Response({'error': 'Request not found'}, status=404)

        if request_obj.status != 'pending':
            return Response({'error': 'Request is not pending'}, status=400)

        old_status = request_obj.status
        request_obj.status = 'approved'
        request_obj.updated_at = timezone.now()
        request_obj.save()

        RequestHistory.objects.create(
            request=request_obj,
            status_from=old_status,
            status_to='approved',
            changed_by=request.user,
            comment=request.data.get('comment', '')
        )
        send_websocket_notification(request_obj, 'approved')
        return Response({'status': 'approved'})


# ==================== REPORT VIEWS ====================
class ReportsViewSet(viewsets.ViewSet):
    """API для формирования отчетов"""
    permission_classes = [IsAuthenticated]

    def _create_excel_response(self, data, headers, filename):
        """Создать Excel файл и вернуть HttpResponse"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Отчет"

        # Заголовки
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Данные
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Автоширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={filename}_{date.today()}.xlsx'
        return response

    def _get_user_role(self, request):
        """Получить роль пользователя"""
        try:
            app_user = AppUser.objects.get(login=request.user.username)
            return app_user.role
        except AppUser.DoesNotExist:
            return 'user'

    @action(detail=False, methods=['get'])
    def all_requests(self, request):
        """Отчет по всем заявкам"""
        role = self._get_user_role(request)
        if role not in ['admin', 'администратор', 'economic_head', 'начальник хоз. отдела']:
            return Response({'error': 'Доступ запрещен'}, status=403)

        requests = Request.objects.all().order_by('-created_at')

        headers = ['№ заявки', 'Цех', 'Заявитель', 'Название', 'Описание', 'Статус', 'Количество', 'Ед. изм.',
                   'Дата создания', 'Комментарий']
        data = []
        for req in requests:
            data.append([
                req.request_number,
                req.shop.title if req.shop else '',
                req.user.login if req.user else '',
                req.title,
                req.description or '',
                req.get_status_display(),
                req.quantity,
                req.unit,
                req.created_at.strftime('%d.%m.%Y %H:%M'),
                req.comment or ''
            ])

        return self._create_excel_response(data, headers, 'all_requests')

    @action(detail=False, methods=['get'], url_path='requests_by_status/(?P<status>[^/.]+)')
    def requests_by_status(self, request, status):
        """Отчет по заявкам с фильтром по статусу"""
        role = self._get_user_role(request)
        if role not in ['admin', 'администратор', 'economic_head', 'начальник хоз. отдела', 'safety_officer',
                        'охрана труда']:
            return Response({'error': 'Доступ запрещен'}, status=403)

        if status != 'all':
            requests = Request.objects.filter(status=status).order_by('-created_at')
        else:
            requests = Request.objects.all().order_by('-created_at')

        headers = ['№ заявки', 'Цех', 'Заявитель', 'Название', 'Описание', 'Статус', 'Количество', 'Ед. изм.',
                   'Дата создания']
        data = []
        for req in requests:
            data.append([
                req.request_number,
                req.shop.title if req.shop else '',
                req.user.login if req.user else '',
                req.title,
                req.description or '',
                req.get_status_display(),
                req.quantity,
                req.unit,
                req.created_at.strftime('%d.%m.%Y %H:%M')
            ])

        return self._create_excel_response(data, headers, f'requests_by_status_{status}')

    @action(detail=False, methods=['get'])
    def requests_by_users(self, request):
        """Отчет по заявкам с группировкой по пользователям"""
        role = self._get_user_role(request)
        if role not in ['admin', 'администратор', 'economic_head', 'начальник хоз. отдела']:
            return Response({'error': 'Доступ запрещен'}, status=403)

        users = AppUser.objects.all().select_related('employee')

        headers = ['Пользователь', 'Сотрудник', 'Цех', 'Количество заявок', 'Заявки (номера)']
        data = []
        for user in users:
            requests = Request.objects.filter(user=user)
            request_numbers = ', '.join([r.request_number for r in requests])
            data.append([
                user.login,
                user.employee.full_name if user.employee else '',
                user.shop.title if user.shop else '',
                requests.count(),
                request_numbers
            ])

        return self._create_excel_response(data, headers, 'requests_by_users')

    @action(detail=False, methods=['get'])
    def pending_safety_requests(self, request):
        """Заявки на рассмотрении в охране труда"""
        role = self._get_user_role(request)
        if role not in ['admin', 'администратор', 'safety_officer', 'охрана труда']:
            return Response({'error': 'Доступ запрещен'}, status=403)

        requests = Request.objects.filter(status='pending').order_by('-created_at')

        headers = ['№ заявки', 'Цех', 'Заявитель', 'Название', 'Описание', 'Статус', 'Дата создания']
        data = []
        for req in requests:
            data.append([
                req.request_number,
                req.shop.title if req.shop else '',
                req.user.login if req.user else '',
                req.title,
                req.description or '',
                req.get_status_display(),
                req.created_at.strftime('%d.%m.%Y %H:%M')
            ])

        return self._create_excel_response(data, headers, 'pending_safety_requests')

    @action(detail=False, methods=['get'])
    def pending_economic_requests(self, request):
        """Заявки на рассмотрении в хоз. отделе"""
        role = self._get_user_role(request)
        if role not in ['admin', 'администратор', 'economic_head', 'начальник хоз. отдела']:
            return Response({'error': 'Доступ запрещен'}, status=403)

        requests = Request.objects.filter(status='hr_approved').order_by('-created_at')

        headers = ['№ заявки', 'Цех', 'Заявитель', 'Название', 'Описание', 'Статус', 'Дата создания']
        data = []
        for req in requests:
            data.append([
                req.request_number,
                req.shop.title if req.shop else '',
                req.user.login if req.user else '',
                req.title,
                req.description or '',
                req.get_status_display(),
                req.created_at.strftime('%d.%m.%Y %H:%M')
            ])

        return self._create_excel_response(data, headers, 'pending_economic_requests')

    @action(detail=False, methods=['get'])
    def my_requests(self, request):
        """Отчет по моим заявкам"""
        try:
            app_user = AppUser.objects.get(login=request.user.username)
            requests = Request.objects.filter(user=app_user).order_by('-created_at')

            headers = ['№ заявки', 'Название', 'Описание', 'Статус', 'Количество', 'Ед. изм.', 'Дата создания',
                       'Комментарий']
            data = []
            for req in requests:
                data.append([
                    req.request_number,
                    req.title,
                    req.description or '',
                    req.get_status_display(),
                    req.quantity,
                    req.unit,
                    req.created_at.strftime('%d.%m.%Y %H:%M'),
                    req.comment or ''
                ])

            return self._create_excel_response(data, headers, 'my_requests')
        except Exception as e:
            return Response({'error': str(e)}, status=400)

    @action(detail=False, methods=['get'], url_path='my_requests_by_status/(?P<status>[^/.]+)')
    def my_requests_by_status(self, request, status):
        """Отчет по моим заявкам с фильтром по статусу"""
        try:
            app_user = AppUser.objects.get(login=request.user.username)
            filter_dict = {'user': app_user}
            if status != 'all':
                filter_dict['status'] = status
            requests = Request.objects.filter(**filter_dict).order_by('-created_at')

            headers = ['№ заявки', 'Название', 'Описание', 'Статус', 'Количество', 'Ед. изм.', 'Дата создания',
                       'Комментарий']
            data = []
            for req in requests:
                data.append([
                    req.request_number,
                    req.title,
                    req.description or '',
                    req.get_status_display(),
                    req.quantity,
                    req.unit,
                    req.created_at.strftime('%d.%m.%Y %H:%M'),
                    req.comment or ''
                ])

            return self._create_excel_response(data, headers, f'my_requests_{status}')
        except Exception as e:
            return Response({'error': str(e)}, status=400)

    @action(detail=False, methods=['get'])
    def my_shop_employees_with_sizes(self, request):
        """Сотрудники цеха с размерами (для нач. цеха и хоз. отдела)"""
        try:
            app_user = AppUser.objects.get(login=request.user.username)
            employees = Employee.objects.filter(shop=app_user.shop, is_active=True).select_related('position')

            headers = ['ФИО', 'Должность', 'Пол', 'Рост (см)', 'Размер одежды', 'Размер обуви', 'Размер головы']
            data = []
            for emp in employees:
                data.append([
                    emp.full_name,
                    emp.position.title,
                    'Мужской' if emp.gender == 'M' else 'Женский' if emp.gender == 'F' else '',
                    emp.heightcm or '',
                    emp.clothing_size or '',
                    emp.shoesize or '',
                    emp.headsize or ''
                ])

            return self._create_excel_response(data, headers, 'my_shop_employees')
        except Exception as e:
            return Response({'error': str(e)}, status=400)

    @action(detail=False, methods=['get'])
    def employees_with_sizes(self, request):
        """Все сотрудники с размерами (для админа)"""
        role = self._get_user_role(request)
        if role not in ['admin', 'администратор']:
            return Response({'error': 'Доступ запрещен'}, status=403)

        employees = Employee.objects.filter(is_active=True).select_related('position', 'shop')

        headers = ['ФИО', 'Должность', 'Цех', 'Пол', 'Рост (см)', 'Размер одежды', 'Размер обуви', 'Размер головы']
        data = []
        for emp in employees:
            data.append([
                emp.full_name,
                emp.position.title,
                emp.shop.title if emp.shop else '',
                'Мужской' if emp.gender == 'M' else 'Женский' if emp.gender == 'F' else '',
                emp.heightcm or '',
                emp.clothing_size or '',
                emp.shoesize or '',
                emp.headsize or ''
            ])

        return self._create_excel_response(data, headers, 'employees_with_sizes')

    @action(detail=False, methods=['get'], url_path='employee_ppe_standards/(?P<employee_id>[^/.]+)')
    def employee_ppe_standards(self, request, employee_id):
        """Нормы выдачи СИЗ по сотруднику"""
        role = self._get_user_role(request)
        if role not in ['admin', 'администратор', 'safety_officer', 'охрана труда']:
            return Response({'error': 'Доступ запрещен'}, status=403)

        try:
            employee = Employee.objects.get(employee_id=employee_id)
            standards = PPEIssueStandard.objects.filter(
                position=employee.position,
                is_active=True
            ).select_related('nomenclature')

            headers = ['СИЗ', 'Ед. изм.', 'Количество', 'Период (мес.)']
            data = []
            for std in standards:
                data.append([
                    std.nomenclature.title,
                    std.nomenclature.unit,
                    float(std.quantity),
                    std.period_months
                ])

            return self._create_excel_response(data, headers, f'employee_{employee_id}_standards')
        except Employee.DoesNotExist:
            return Response({'error': 'Сотрудник не найден'}, status=404)

    @action(detail=False, methods=['get'])
    def position_ppe_standards(self, request):
        """Нормы выдачи СИЗ по должностям"""
        role = self._get_user_role(request)
        if role not in ['admin', 'администратор', 'safety_officer', 'охрана труда']:
            return Response({'error': 'Доступ запрещен'}, status=403)

        positions = Position.objects.all()

        headers = ['Должность', 'СИЗ', 'Ед. изм.', 'Количество', 'Период (мес.)']
        data = []
        for position in positions:
            standards = PPEIssueStandard.objects.filter(
                position=position,
                is_active=True
            ).select_related('nomenclature')
            for std in standards:
                data.append([
                    position.title,
                    std.nomenclature.title,
                    std.nomenclature.unit,
                    float(std.quantity),
                    std.period_months
                ])

        return self._create_excel_response(data, headers, 'position_ppe_standards')

    @action(detail=False, methods=['get'])
    def all_ppe_issues(self, request):
        """Все выдачи СИЗ"""
        role = self._get_user_role(request)
        if role not in ['admin', 'администратор', 'economic_head', 'начальник хоз. отдела']:
            return Response({'error': 'Доступ запрещен'}, status=403)

        issues = PPEIssueRecord.objects.select_related('employee', 'nomenclature', 'issued_by').order_by('-issue_date')

        headers = ['Сотрудник', 'Должность', 'Цех', 'СИЗ', 'Размер', 'Количество', 'Ед. изм.', 'Дата выдачи',
                   'След. выдача', 'Кто выдал']
        data = []
        for issue in issues:
            data.append([
                issue.employee.full_name,
                issue.employee.position.title,
                issue.employee.shop.title if issue.employee.shop else '',
                issue.nomenclature.title,
                issue.size or '',
                float(issue.quantity),
                issue.nomenclature.unit,
                issue.issue_date.strftime('%d.%m.%Y'),
                issue.next_issue_date.strftime('%d.%m.%Y') if issue.next_issue_date else '',
                issue.issued_by.login if issue.issued_by else ''
            ])

        return self._create_excel_response(data, headers, 'all_ppe_issues')

    @action(detail=False, methods=['get'], url_path='employee_ppe_issues/(?P<employee_id>[^/.]+)')
    def employee_ppe_issues(self, request, employee_id):
        """Выдачи СИЗ по сотруднику"""
        role = self._get_user_role(request)
        if role not in ['admin', 'администратор', 'economic_head', 'начальник хоз. отдела']:
            return Response({'error': 'Доступ запрещен'}, status=403)

        try:
            issues = PPEIssueRecord.objects.filter(
                employee_id=employee_id
            ).select_related('nomenclature', 'issued_by').order_by('-issue_date')

            headers = ['СИЗ', 'Размер', 'Количество', 'Ед. изм.', 'Дата выдачи', 'След. выдача', 'Кто выдал']
            data = []
            for issue in issues:
                data.append([
                    issue.nomenclature.title,
                    issue.size or '',
                    float(issue.quantity),
                    issue.nomenclature.unit,
                    issue.issue_date.strftime('%d.%m.%Y'),
                    issue.next_issue_date.strftime('%d.%m.%Y') if issue.next_issue_date else '',
                    issue.issued_by.login if issue.issued_by else ''
                ])

            return self._create_excel_response(data, headers, f'employee_{employee_id}_issues')
        except Exception as e:
            return Response({'error': str(e)}, status=400)

    @action(detail=False, methods=['get'])
    def mass_issue_report(self, request):
        """Отчет по массовой выдаче"""
        role = self._get_user_role(request)
        if role not in ['admin', 'администратор', 'economic_head', 'начальник хоз. отдела']:
            return Response({'error': 'Доступ запрещен'}, status=403)

        # Здесь можно добавить логику для отчета по массовой выдаче
        # Пока возвращаем заглушку
        headers = ['Дата', 'Сотрудник', 'СИЗ', 'Количество']
        data = [['Нет данных', '', '', '']]
        return self._create_excel_response(data, headers, 'mass_issue_report')

    @action(detail=False, methods=['get'])
    def users_report(self, request):
        """Отчет по пользователям системы"""
        role = self._get_user_role(request)
        if role not in ['admin', 'администратор']:
            return Response({'error': 'Доступ запрещен'}, status=403)

        users = AppUser.objects.select_related('employee', 'shop')

        headers = ['Логин', 'Роль', 'Сотрудник', 'Цех']
        data = []
        for user in users:
            data.append([
                user.login,
                user.role,
                user.employee.full_name if user.employee else '',
                user.shop.title if user.shop else ''
            ])

        return self._create_excel_response(data, headers, 'users_report')

    @action(detail=False, methods=['get'])
    def shops_report(self, request):
        """Отчет по цехам"""
        role = self._get_user_role(request)
        if role not in ['admin', 'администратор']:
            return Response({'error': 'Доступ запрещен'}, status=403)

        shops = Shop.objects.all()

        headers = ['Код', 'Название']
        data = []
        for shop in shops:
            data.append([
                shop.code,
                shop.title
            ])

        return self._create_excel_response(data, headers, 'shops_report')

    @action(detail=False, methods=['get'])
    def positions_report(self, request):
        """Отчет по должностям"""
        role = self._get_user_role(request)
        if role not in ['admin', 'администратор']:
            return Response({'error': 'Доступ запрещен'}, status=403)

        positions = Position.objects.all()

        headers = ['Должность']
        data = []
        for pos in positions:
            data.append([pos.title])

        return self._create_excel_response(data, headers, 'positions_report')

    @action(detail=False, methods=['get'])
    def sizes_report(self, request):
        """Отчет по ГОСТ размерам"""
        role = self._get_user_role(request)
        if role not in ['admin', 'администратор']:
            return Response({'error': 'Доступ запрещен'}, status=403)

        from .models import ClothingSizeGOST, FootwearSizeGOST, HeadwearSizeGOST

        wb = Workbook()

        # Лист с размерами одежды
        ws1 = wb.active
        ws1.title = "Размеры одежды"
        headers1 = ['Код', 'Пол', 'Рост от', 'Рост до', 'Грудь', 'Талия', 'Бедра']
        for col, header in enumerate(headers1, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        row = 2
        for size in ClothingSizeGOST.objects.all().order_by('sort_order'):
            ws1.cell(row=row, column=1, value=size.size_code)
            ws1.cell(row=row, column=2, value='М' if size.gender == 'M' else 'Ж' if size.gender == 'F' else 'У')
            ws1.cell(row=row, column=3, value=size.height_min or '')
            ws1.cell(row=row, column=4, value=size.height_max or '')
            ws1.cell(row=row, column=5, value=size.chest_circumference or '')
            ws1.cell(row=row, column=6, value=size.waist_circumference or '')
            ws1.cell(row=row, column=7, value=size.hip_circumference or '')
            row += 1

        # Лист с размерами обуви
        ws2 = wb.create_sheet("Размеры обуви")
        headers2 = ['Рос. размер', 'EU', 'US', 'Длина стопы от', 'Длина стопы до']
        for col, header in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        row = 2
        for size in FootwearSizeGOST.objects.all().order_by('sort_order'):
            ws2.cell(row=row, column=1, value=size.size_ru)
            ws2.cell(row=row, column=2, value=size.size_eu or '')
            ws2.cell(row=row, column=3, value=size.size_us or '')
            ws2.cell(row=row, column=4, value=size.foot_length_min or '')
            ws2.cell(row=row, column=5, value=size.foot_length_max or '')
            row += 1

        # Лист с размерами головы
        ws3 = wb.create_sheet("Размеры головы")
        headers3 = ['Код размера', 'Обхват головы от', 'Обхват головы до']
        for col, header in enumerate(headers3, 1):
            cell = ws3.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        row = 2
        for size in HeadwearSizeGOST.objects.all().order_by('sort_order'):
            ws3.cell(row=row, column=1, value=size.size_code)
            ws3.cell(row=row, column=2, value=size.head_circumference_min or '')
            ws3.cell(row=row, column=3, value=size.head_circumference_max or '')
            row += 1

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=sizes_report_{date.today()}.xlsx'
        return response

class EconomicReportViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=['get'])
    def generate_consolidated_report(self, request):
        approved_requests = Request.objects.filter(status='approved').order_by('created_at')
        if not approved_requests.exists():
            return Response({'error': 'No approved requests found'}, status=404)

        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Все заявки по сотрудникам"

        headers = ['№ заявки', 'Цех', 'ФИО сотрудника', 'Должность', 'СИЗ', 'Размер', 'Количество', 'Ед. изм.']
        for col, header in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        row = 2
        for req in approved_requests:
            request_employees = RequestEmployee.objects.filter(request=req)
            for req_emp in request_employees:
                items = RequestEmployeeItem.objects.filter(request_employee=req_emp)
                for item in items:
                    ws1.cell(row=row, column=1, value=req.request_number)
                    ws1.cell(row=row, column=2, value=req.shop.title)
                    ws1.cell(row=row, column=3, value=req_emp.employee.full_name)
                    ws1.cell(row=row, column=4, value=req_emp.employee.position.title)
                    ws1.cell(row=row, column=5, value=item.nomenclature.title)
                    ws1.cell(row=row, column=6, value=item.size)
                    ws1.cell(row=row, column=7, value=float(item.quantity))
                    ws1.cell(row=row, column=8, value=item.nomenclature.unit)
                    row += 1

        ws2 = wb.create_sheet("Для поставщика")
        supplier_headers = ['СИЗ', 'Размер', 'Общее количество', 'Ед. изм.']
        for col, header in enumerate(supplier_headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        summary = {}
        for req in approved_requests:
            request_employees = RequestEmployee.objects.filter(request=req)
            for req_emp in request_employees:
                items = RequestEmployeeItem.objects.filter(request_employee=req_emp)
                for item in items:
                    key = (item.nomenclature.title, item.size)
                    if key not in summary:
                        summary[key] = {'quantity': 0, 'unit': item.nomenclature.unit}
                    summary[key]['quantity'] += float(item.quantity)

        row = 2
        for (name, size), data in summary.items():
            ws2.cell(row=row, column=1, value=name)
            ws2.cell(row=row, column=2, value=size)
            ws2.cell(row=row, column=3, value=data['quantity'])
            ws2.cell(row=row, column=4, value=data['unit'])
            row += 1

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=consolidated_report_{date.today()}.xlsx'
        return response


# ==================== ADMIN VIEWS ====================

class AdminUserViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated, IsAdminUserPermission]

    @action(detail=False, methods=['get'])
    def get_all_users(self, request):
        app_users = AppUser.objects.select_related('employee', 'shop').all()
        result = []
        for au in app_users:
            result.append({
                'user_id': au.user_id,
                'login': au.login,
                'role': au.role,
                'employee_id': au.employee.employee_id if au.employee else None,
                'employee_name': au.employee.full_name if au.employee else None,
                'shop_id': au.shop.shop_id if au.shop else None,
                'shop_name': au.shop.title if au.shop else None,
            })
        return Response(result)

    @action(detail=False, methods=['post'])
    def create_user(self, request):
        login = request.data.get('login')
        password = request.data.get('password')
        role = request.data.get('role')
        employee_id = request.data.get('employee_id')
        shop_id = request.data.get('shop_id')

        if not login or not password:
            return Response({'error': 'Login and password required'}, status=400)
        if AppUser.objects.filter(login=login).exists():
            return Response({'error': 'User already exists'}, status=400)

        try:
            employee = Employee.objects.get(employee_id=employee_id) if employee_id else None
            shop = Shop.objects.get(shop_id=shop_id) if shop_id else None
        except (Employee.DoesNotExist, Shop.DoesNotExist):
            return Response({'error': 'Employee or Shop not found'}, status=404)

        app_user = AppUser.objects.create(
            login=login,
            password=make_password(password),
            role=role,
            employee=employee,
            shop=shop
        )

        django_user, _ = User.objects.get_or_create(
            username=login,
            defaults={
                'email': f"{login}@esmsystem.ru",
                'is_active': True,
                'is_staff': role == 'admin',
                'is_superuser': role == 'admin',
            }
        )
        return Response({'message': 'User created successfully', 'user_id': app_user.user_id})

    @action(detail=False, methods=['put'])
    def update_user(self, request):  # Убедитесь, что это PUT метод
        """Обновить данные пользователя"""
        user_id = request.data.get('user_id')
        role = request.data.get('role')
        employee_id = request.data.get('employee_id')
        shop_id = request.data.get('shop_id')
        password = request.data.get('password')

        try:
            app_user = AppUser.objects.get(user_id=user_id)
        except AppUser.DoesNotExist:
            return Response({'error': 'User not found'}, status=404)

        if role:
            app_user.role = role
        if employee_id:
            try:
                app_user.employee = Employee.objects.get(employee_id=employee_id)
            except Employee.DoesNotExist:
                pass
        if shop_id:
            try:
                app_user.shop = Shop.objects.get(shop_id=shop_id)
            except Shop.DoesNotExist:
                pass
        if password:
            from django.contrib.auth.hashers import make_password
            app_user.password = make_password(password)

        app_user.save()

        # Обновляем Django пользователя
        from django.contrib.auth.models import User
        try:
            django_user = User.objects.get(username=app_user.login)
            django_user.is_staff = role == 'admin'
            django_user.is_superuser = role == 'admin'
            if password:
                django_user.set_password(password)
            django_user.save()
        except User.DoesNotExist:
            pass

        return Response({'message': 'User updated successfully'})
    @action(detail=False, methods=['delete'])
    def delete_user(self, request):
        user_id = request.data.get('user_id')
        try:
            app_user = AppUser.objects.get(user_id=user_id)
            login = app_user.login
            app_user.delete()
            User.objects.filter(username=login).delete()
            return Response({'message': 'User deleted successfully'})
        except AppUser.DoesNotExist:
            return Response({'error': 'User not found'}, status=404)


class AdminShopViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated, IsAdminUserPermission]

    @action(detail=False, methods=['get'])
    def get_all_shops(self, request):
        shops = Shop.objects.all()
        return Response([{'shop_id': s.shop_id, 'title': s.title, 'code': s.code} for s in shops])

    @action(detail=False, methods=['post'])
    def create_shop(self, request):
        title = request.data.get('title')
        code = request.data.get('code')
        if not title or not code:
            return Response({'error': 'Title and code required'}, status=400)
        shop = Shop.objects.create(title=title, code=code)
        return Response({'message': 'Shop created', 'shop_id': shop.shop_id})

    @action(detail=False, methods=['put'])
    def update_shop(self, request):
        shop_id = request.data.get('shop_id')
        title = request.data.get('title')
        code = request.data.get('code')
        try:
            shop = Shop.objects.get(shop_id=shop_id)
            if title:
                shop.title = title
            if code:
                shop.code = code
            shop.save()
            return Response({'message': 'Shop updated'})
        except Shop.DoesNotExist:
            return Response({'error': 'Shop not found'}, status=404)

    @action(detail=False, methods=['delete'])
    def delete_shop(self, request):
        shop_id = request.data.get('shop_id')
        try:
            Shop.objects.get(shop_id=shop_id).delete()
            return Response({'message': 'Shop deleted'})
        except Shop.DoesNotExist:
            return Response({'error': 'Shop not found'}, status=404)


class AdminEmployeeViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated, IsAdminUserPermission]

    @action(detail=False, methods=['get'])
    def get_all_employees(self, request):
        employees = Employee.objects.select_related('shop', 'position').all()
        result = []
        for emp in employees:
            result.append({
                'employee_id': emp.employee_id,
                'full_name': emp.full_name,
                'first_name': emp.first_name,
                'second_name': emp.second_name,
                'last_name': emp.last_name,
                'position_id': emp.position.position_id,
                'position_name': emp.position.title,
                'shop_id': emp.shop.shop_id,
                'shop_name': emp.shop.title,
                'is_active': emp.is_active,
                'hire_date': emp.hire_date,
                'gender': emp.gender,
                'heightcm': emp.heightcm,
                'clothing_size': emp.clothing_size,
                'shoesize': emp.shoesize,
                'headsize': emp.headsize,
            })
        return Response(result)

    @action(detail=False, methods=['get'])
    def get_all_employees(self, request):
        """Получить ВСЕХ сотрудников (только для администраторов)"""
        employees = Employee.objects.select_related('shop', 'position').filter(is_active=True)
        result = []
        for emp in employees:
            result.append({
                'employee_id': emp.employee_id,
                'full_name': emp.full_name,
                'position_name': emp.position.title,
                'gender': emp.gender,
                'heightcm': emp.heightcm,
                'clothing_size': emp.clothing_size,
                'shoesize': emp.shoesize,
                'headsize': emp.headsize,
                'shop_name': emp.shop.title if emp.shop else None,
            })
        return Response(result)

    @action(detail=False, methods=['post'])
    def create_employee(self, request):
        """Создать нового сотрудника"""
        data = request.data

        # Проверка обязательных полей
        if not data.get('last_name'):
            return Response({'error': 'Фамилия обязательна'}, status=400)
        if not data.get('first_name'):
            return Response({'error': 'Имя обязательно'}, status=400)
        if not data.get('position_id'):
            return Response({'error': 'Должность обязательна'}, status=400)
        if not data.get('shop_id'):
            return Response({'error': 'Цех обязателен'}, status=400)

        try:
            shop = Shop.objects.get(shop_id=data.get('shop_id'))
            position = Position.objects.get(position_id=data.get('position_id'))
        except (Shop.DoesNotExist, Position.DoesNotExist) as e:
            return Response({'error': f'Цех или должность не найдены: {str(e)}'}, status=404)

        # Функция для преобразования чисел
        def parse_int_or_none(value):
            if value is None or value == '':
                return None
            try:
                return int(value)
            except (ValueError, TypeError):
                return None

        try:
            # НЕ указываем employee_id - пусть база данных сама назначит следующий ID
            employee = Employee.objects.create(
                shop=shop,
                position=position,
                first_name=data.get('first_name', ''),
                second_name=data.get('second_name', ''),
                last_name=data.get('last_name', ''),
                is_active=True,
                hire_date=date.today(),
                gender=data.get('gender') if data.get('gender') else None,
                heightcm=parse_int_or_none(data.get('heightcm')),
                clothing_size=data.get('clothing_size') if data.get('clothing_size') else None,
                shoesize=parse_int_or_none(data.get('shoesize')),
                headsize=parse_int_or_none(data.get('headsize')),
            )
            return Response({'message': 'Employee created', 'employee_id': employee.employee_id})
        except Exception as e:
            return Response({'error': str(e)}, status=400)

    @action(detail=False, methods=['put'])
    def update_employee(self, request):
        """Обновить данные сотрудника"""
        employee_id = request.data.get('employee_id')
        try:
            employee = Employee.objects.get(employee_id=employee_id)
        except Employee.DoesNotExist:
            return Response({'error': 'Employee not found'}, status=404)

        # Обновляем поля с проверкой на None
        if 'first_name' in request.data and request.data['first_name'] is not None:
            employee.first_name = request.data['first_name']
        if 'second_name' in request.data and request.data['second_name'] is not None:
            employee.second_name = request.data['second_name']
        if 'last_name' in request.data and request.data['last_name'] is not None:
            employee.last_name = request.data['last_name']
        if 'gender' in request.data:
            employee.gender = request.data['gender'] if request.data['gender'] else None
        if 'clothing_size' in request.data:
            employee.clothing_size = request.data['clothing_size'] if request.data['clothing_size'] else None

        # Числовые поля - проверяем на None и пустые строки
        if 'heightcm' in request.data:
            height_val = request.data['heightcm']
            employee.heightcm = int(height_val) if height_val and height_val != '' else None

        if 'shoesize' in request.data:
            shoe_val = request.data['shoesize']
            employee.shoesize = int(shoe_val) if shoe_val and shoe_val != '' else None

        if 'headsize' in request.data:
            head_val = request.data['headsize']
            employee.headsize = int(head_val) if head_val and head_val != '' else None

        if 'is_active' in request.data:
            employee.is_active = request.data['is_active']

        if 'position_id' in request.data and request.data['position_id']:
            try:
                employee.position = Position.objects.get(position_id=request.data['position_id'])
            except Position.DoesNotExist:
                pass

        if 'shop_id' in request.data and request.data['shop_id']:
            try:
                employee.shop = Shop.objects.get(shop_id=request.data['shop_id'])
            except Shop.DoesNotExist:
                pass

        employee.save()
        return Response({'message': 'Employee updated'})

    @action(detail=False, methods=['delete'])
    def delete_employee(self, request):
        employee_id = request.data.get('employee_id')
        try:
            Employee.objects.get(employee_id=employee_id).delete()
            return Response({'message': 'Employee deleted'})
        except Employee.DoesNotExist:
            return Response({'error': 'Employee not found'}, status=404)


class AdminPositionViewSet(viewsets.ViewSet):
    """API для управления должностями"""
    permission_classes = [IsAuthenticated]  # Измените с IsAdminUserPermission на IsAuthenticated

    @action(detail=False, methods=['get'])
    def get_all_positions(self, request):
        """Получить все должности"""
        positions = Position.objects.all()
        return Response([{'position_id': p.position_id, 'title': p.title} for p in positions])
    @action(detail=False, methods=['post'])
    def create_position(self, request):
        title = request.data.get('title')
        if not title:
            return Response({'error': 'Title required'}, status=400)
        position = Position.objects.create(title=title)
        return Response({'message': 'Position created', 'position_id': position.position_id})

    @action(detail=False, methods=['put'])
    def update_position(self, request):
        position_id = request.data.get('position_id')
        title = request.data.get('title')
        try:
            position = Position.objects.get(position_id=position_id)
            if title:
                position.title = title
            position.save()
            return Response({'message': 'Position updated'})
        except Position.DoesNotExist:
            return Response({'error': 'Position not found'}, status=404)

    @action(detail=False, methods=['delete'])
    def delete_position(self, request):
        position_id = request.data.get('position_id')
        try:
            Position.objects.get(position_id=position_id).delete()
            return Response({'message': 'Position deleted'})
        except Position.DoesNotExist:
            return Response({'error': 'Position not found'}, status=404)

class IsSafetyOfficer(permissions.BasePermission):
    """Проверка, является ли пользователь сотрудником охраны труда"""
    def has_permission(self, request, view):
        if not request.user.is_authenticated:
            return False
        try:
            app_user = AppUser.objects.get(login=request.user.username)
            return app_user.role == 'safety_officer' or request.user.is_superuser
        except AppUser.DoesNotExist:
            return request.user.is_superuser


class IsSafetyOrAdmin(permissions.BasePermission):
    """Проверка, является ли пользователь сотрудником охраны труда или администратором"""
    def has_permission(self, request, view):
        return request.user.is_authenticated and (
            request.user.is_superuser or
            request.user.groups.filter(name='safety_officer').exists()
        )


class SafetyStandardViewSet(viewsets.ViewSet):
    """API для управления нормами выдачи (охрана труда)"""
    permission_classes = [IsAuthenticated]

    def list(self, request):
        """Получить все нормы выдачи"""
        from .models import PPEIssueStandard

        standards = PPEIssueStandard.objects.select_related('position', 'nomenclature').all()
        result = []
        for std in standards:
            result.append({
                'standard_id': std.standard_id,
                'position_id': std.position.position_id,
                'position_title': std.position.title,
                'nomenclature_id': std.nomenclature.nomenclature_id,
                'nomenclature_title': std.nomenclature.title,
                'unit': std.nomenclature.unit,
                'quantity': float(std.quantity),
                'period_months': std.period_months,
                'is_active': std.is_active,
            })
        return Response(result)

    @action(detail=False, methods=['post'])
    def create_nomenclature(self, request):
        """Создать новый СИЗ (номенклатуру)"""
        title = request.data.get('title')
        unit = request.data.get('unit', 'шт')
        shelf_life_months = request.data.get('shelf_life_months', 12)

        if not title:
            return Response({'error': 'Название СИЗ обязательно'}, status=400)

        # Проверяем, не существует ли уже такой СИЗ
        if Nomenclature.objects.filter(title__iexact=title).exists():
            return Response({'error': 'СИЗ с таким названием уже существует'}, status=400)

        nomenclature = Nomenclature.objects.create(
            title=title,
            unit=unit,
            shelf_life_months=shelf_life_months,
            is_active=True
        )

        return Response({
            'message': 'СИЗ создан',
            'nomenclature_id': nomenclature.nomenclature_id,
            'title': nomenclature.title,
            'unit': nomenclature.unit,
            'shelf_life_months': nomenclature.shelf_life_months
        }, status=201)

    @action(detail=False, methods=['get'])
    def get_nomenclatures(self, request):
        """Получить все активные СИЗ (с возможностью создания)"""
        nomenclatures = Nomenclature.objects.filter(is_active=True).order_by('title')
        result = [{
            'nomenclature_id': n.nomenclature_id,
            'title': n.title,
            'unit': n.unit,
            'shelf_life_months': n.shelf_life_months,
        } for n in nomenclatures]
        return Response(result)
    def list(self, request):
        """Получить все нормы выдачи (GET /api/safety/standards/)"""
        from .models import PPEIssueStandard

        standards = PPEIssueStandard.objects.select_related('position', 'nomenclature').all()
        result = []
        for std in standards:
            result.append({
                'standard_id': std.standard_id,
                'position_id': std.position.position_id,
                'position_title': std.position.title,
                'nomenclature_id': std.nomenclature.nomenclature_id,
                'nomenclature_title': std.nomenclature.title,
                'unit': std.nomenclature.unit,
                'quantity': float(std.quantity),
                'period_months': std.period_months,
                'is_active': std.is_active,
            })
        return Response(result)

    @action(detail=False, methods=['get'])
    def get_all_standards(self, request):
        """Альтернативный метод для получения норм"""
        return self.list(request)
    @action(detail=False, methods=['get'])
    def get_standards_by_position(self, request):
        """Получить нормы для конкретной должности"""
        position_id = request.query_params.get('position_id')
        if not position_id:
            return Response({'error': 'position_id required'}, status=400)

        standards = PPEIssueStandard.objects.filter(
            position_id=position_id,
            is_active=True
        ).select_related('nomenclature')

        result = []
        for std in standards:
            result.append({
                'standard_id': std.standard_id,
                'nomenclature_id': std.nomenclature.nomenclature_id,
                'nomenclature_title': std.nomenclature.title,
                'unit': std.nomenclature.unit,
                'quantity': float(std.quantity),
                'period_months': std.period_months,
            })
        return Response(result)

    @action(detail=False, methods=['post'])
    def create_standard(self, request):
        """Создать новую норму выдачи"""
        position_id = request.data.get('position_id')
        nomenclature_id = request.data.get('nomenclature_id')
        quantity = request.data.get('quantity', 1)
        period_months = request.data.get('period_months', 12)

        if not position_id or not nomenclature_id:
            return Response({'error': 'position_id and nomenclature_id required'}, status=400)

        try:
            position = Position.objects.get(position_id=position_id)
            nomenclature = Nomenclature.objects.get(nomenclature_id=nomenclature_id)
        except (Position.DoesNotExist, Nomenclature.DoesNotExist):
            return Response({'error': 'Position or Nomenclature not found'}, status=404)

        standard, created = PPEIssueStandard.objects.get_or_create(
            position=position,
            nomenclature=nomenclature,
            defaults={
                'quantity': quantity,
                'period_months': period_months,
                'is_active': True
            }
        )

        if not created:
            standard.quantity = quantity
            standard.period_months = period_months
            standard.is_active = True
            standard.save()

        return Response({'message': 'Standard saved', 'standard_id': standard.standard_id})

    @action(detail=False, methods=['put'])
    def update_standard(self, request):
        """Обновить норму выдачи"""
        standard_id = request.data.get('standard_id')
        quantity = request.data.get('quantity')
        period_months = request.data.get('period_months')
        is_active = request.data.get('is_active')

        try:
            standard = PPEIssueStandard.objects.get(standard_id=standard_id)
            if quantity is not None:
                standard.quantity = quantity
            if period_months is not None:
                standard.period_months = period_months
            if is_active is not None:
                standard.is_active = is_active
            standard.save()
            return Response({'message': 'Standard updated'})
        except PPEIssueStandard.DoesNotExist:
            return Response({'error': 'Standard not found'}, status=404)

    @action(detail=False, methods=['delete'])
    def delete_standard(self, request):
        """Удалить норму выдачи"""
        standard_id = request.data.get('standard_id')
        try:
            PPEIssueStandard.objects.get(standard_id=standard_id).delete()
            return Response({'message': 'Standard deleted'})
        except PPEIssueStandard.DoesNotExist:
            return Response({'error': 'Standard not found'}, status=404)

class NomenclatureAdminViewSet(viewsets.ViewSet):
    """API для управления справочником СИЗ"""
    # Разрешаем просмотр всем аутентифицированным пользователям
    # Но изменение только для админов и охраны труда
    permission_classes = [IsAuthenticated]

    def get_permissions(self):
        # Для методов изменения (create, update, destroy) используем более строгие права
        if self.action in ['create', 'update', 'destroy']:
            return [IsAuthenticated(), IsSafetyOrAdmin()]
        return [IsAuthenticated()]

    def list(self, request):
        """GET /api/safety/nomenclatures/ - получить список СИЗ (доступно всем)"""
        nomenclatures = Nomenclature.objects.filter(is_active=True).order_by('title')
        result = [{
            'nomenclature_id': n.nomenclature_id,
            'title': n.title,
            'unit': n.unit,
            'shelf_life_months': n.shelf_life_months,
            'is_active': n.is_active,
        } for n in nomenclatures]
        return Response(result)

    def create(self, request):
        """POST /api/safety/nomenclatures/ - создать новый СИЗ (только админ/охрана)"""
        self.check_object_permissions(request, None)
        title = request.data.get('title')
        unit = request.data.get('unit', 'шт')
        shelf_life_months = request.data.get('shelf_life_months', 12)

        if not title:
            return Response({'error': 'Название СИЗ обязательно'}, status=400)

        if Nomenclature.objects.filter(title__iexact=title).exists():
            return Response({'error': 'СИЗ с таким названием уже существует'}, status=400)

        nomenclature = Nomenclature.objects.create(
            title=title,
            unit=unit,
            shelf_life_months=shelf_life_months,
            is_active=True
        )

        return Response({
            'message': 'СИЗ успешно создан',
            'nomenclature_id': nomenclature.nomenclature_id,
            'title': nomenclature.title,
            'unit': nomenclature.unit,
            'shelf_life_months': nomenclature.shelf_life_months,
        }, status=201)

    def update(self, request, pk=None):
        """PUT /api/safety/nomenclatures/{id}/ - обновить СИЗ (только админ/охрана)"""
        self.check_object_permissions(request, None)
        try:
            nomenclature = Nomenclature.objects.get(pk=pk)
        except Nomenclature.DoesNotExist:
            return Response({'error': 'СИЗ не найден'}, status=404)

        title = request.data.get('title')
        unit = request.data.get('unit')
        shelf_life_months = request.data.get('shelf_life_months')

        # Если название не меняется - пропускаем проверку уникальности
        if title and title.lower() != nomenclature.title.lower():
            # Проверяем, существует ли СИЗ с таким названием (исключая текущий)
            if Nomenclature.objects.filter(title__iexact=title).exclude(pk=pk).exists():
                return Response({'error': 'СИЗ с таким названием уже существует'}, status=400)
            nomenclature.title = title

        if unit:
            nomenclature.unit = unit
        if shelf_life_months is not None:
            nomenclature.shelf_life_months = shelf_life_months

        nomenclature.save()
        return Response({'message': 'СИЗ обновлен', 'nomenclature': {
            'nomenclature_id': nomenclature.nomenclature_id,
            'title': nomenclature.title,
            'unit': nomenclature.unit,
            'shelf_life_months': nomenclature.shelf_life_months,
            'is_active': nomenclature.is_active,
        }})
    def destroy(self, request, pk=None):
        """DELETE /api/safety/nomenclatures/{id}/ - деактивировать СИЗ (только админ/охрана)"""
        self.check_object_permissions(request, None)
        try:
            nomenclature = Nomenclature.objects.get(pk=pk)
            nomenclature.is_active = False
            nomenclature.save()
            return Response({'message': 'СИЗ деактивирован'})
        except Nomenclature.DoesNotExist:
            return Response({'error': 'СИЗ не найден'}, status=404)


class PPEIssueViewSet(viewsets.ViewSet):
    """API для управления выдачами СИЗ"""
    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=['get'])
    def get_all_issues(self, request):
        """Получить все выдачи СИЗ с информацией о сроках"""
        try:
            # Проверяем права (только admin и economic_head)
            app_user = AppUser.objects.get(login=request.user.username)
            if not (request.user.is_superuser or app_user.role in ['admin', 'economic_head']):
                return Response({'error': 'Доступ запрещен'}, status=403)

            issues = PPEIssueRecord.objects.select_related(
                'employee', 'nomenclature', 'issued_by'
            ).order_by('-issue_date')

            today = date.today()
            result = []

            for issue in issues:
                # Рассчитываем дни до следующей выдачи
                days_until_next = (issue.next_issue_date - today).days if issue.next_issue_date else 365

                # Определяем статус и цвет
                if days_until_next <= 0:
                    status = 'expired'
                    color = '#f44336'
                    status_text = 'Требуется выдача'
                elif days_until_next <= 60:  # менее 2 месяцев
                    status = 'critical'
                    color = '#f44336'
                    status_text = f'Срочно! Осталось {days_until_next} дн.'
                elif days_until_next <= 150:  # до 5 месяцев
                    status = 'warning'
                    color = '#ff9800'
                    status_text = f'Скоро замена, осталось {days_until_next} дн.'
                else:
                    status = 'good'
                    color = '#4caf50'
                    status_text = f'Срок не истек, осталось {days_until_next} дн.'

                result.append({
                    'issue_id': issue.issue_id,
                    'employee_id': issue.employee.employee_id,
                    'employee_name': issue.employee.full_name,
                    'position_name': issue.employee.position.title,
                    'shop_name': issue.employee.shop.title if issue.employee.shop else '',
                    'nomenclature_id': issue.nomenclature.nomenclature_id,
                    'nomenclature_title': issue.nomenclature.title,
                    'unit': issue.nomenclature.unit,
                    'size': issue.size,
                    'quantity': float(issue.quantity),
                    'issue_date': issue.issue_date.isoformat(),
                    'next_issue_date': issue.next_issue_date.isoformat() if issue.next_issue_date else None,
                    'days_until_next': days_until_next,
                    'status': status,
                    'color': color,
                    'status_text': status_text,
                    'issued_by': issue.issued_by.login if issue.issued_by else 'Система',
                    'comment': issue.comment or '',
                })

            return Response(result)
        except AppUser.DoesNotExist:
            return Response({'error': 'User not found'}, status=404)

    @action(detail=False, methods=['get'])
    def get_employee_issues(self, request):
        """Получить выдачи СИЗ для конкретного сотрудника"""
        employee_id = request.query_params.get('employee_id')
        if not employee_id:
            return Response({'error': 'employee_id required'}, status=400)

        issues = PPEIssueRecord.objects.filter(
            employee_id=employee_id
        ).select_related('nomenclature').order_by('-issue_date')

        today = date.today()
        result = []

        for issue in issues:
            days_until_next = (issue.next_issue_date - today).days if issue.next_issue_date else 365

            result.append({
                'issue_id': issue.issue_id,
                'nomenclature_title': issue.nomenclature.title,
                'size': issue.size,
                'quantity': float(issue.quantity),
                'unit': issue.nomenclature.unit,
                'issue_date': issue.issue_date.isoformat(),
                'next_issue_date': issue.next_issue_date.isoformat() if issue.next_issue_date else None,
                'days_until_next': days_until_next,
            })

        return Response(result)

    @action(detail=False, methods=['post'])
    def create_issue(self, request):
        """Записать новую выдачу СИЗ"""
        employee_id = request.data.get('employee_id')
        nomenclature_id = request.data.get('nomenclature_id')
        issue_date = request.data.get('issue_date')
        size = request.data.get('size')
        quantity = request.data.get('quantity', 1)
        period_months = request.data.get('period_months', 12)
        comment = request.data.get('comment', '')

        if not employee_id or not nomenclature_id or not issue_date:
            return Response({'error': 'employee_id, nomenclature_id and issue_date required'}, status=400)

        try:
            employee = Employee.objects.get(employee_id=employee_id)
            nomenclature = Nomenclature.objects.get(nomenclature_id=nomenclature_id)
            app_user = AppUser.objects.get(login=request.user.username)
        except (Employee.DoesNotExist, Nomenclature.DoesNotExist, AppUser.DoesNotExist):
            return Response({'error': 'Employee, Nomenclature or User not found'}, status=404)

        from dateutil.relativedelta import relativedelta
        issue_date_obj = datetime.strptime(issue_date, '%Y-%m-%d').date()
        next_issue_date = issue_date_obj + relativedelta(months=period_months)

        issue = PPEIssueRecord.objects.create(
            employee=employee,
            nomenclature=nomenclature,
            issue_date=issue_date_obj,
            size=size,
            quantity=quantity,
            next_issue_date=next_issue_date,
            issued_by=app_user,
            comment=comment
        )

        return Response({
            'message': 'Выдача записана',
            'issue_id': issue.issue_id,
            'next_issue_date': next_issue_date.isoformat()
        }, status=201)
class IsSafetyOrAdmin(permissions.BasePermission):
    """Проверка, является ли пользователь сотрудником охраны труда или администратором"""
    def has_permission(self, request, view):
        if not request.user.is_authenticated:
            return False
        # Проверяем роль через AppUser
        try:
            app_user = AppUser.objects.get(login=request.user.username)
            return app_user.role == 'safety_officer' or request.user.is_superuser
        except AppUser.DoesNotExist:
            return request.user.is_superuser

class NomenclatureViewSet(viewsets.ReadOnlyModelViewSet):
    """API для получения списка СИЗ (номенклатуры)"""
    permission_classes = [IsAuthenticated]
    queryset = Nomenclature.objects.filter(is_active=True).order_by('title')
    serializer_class = NomenclatureSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['title']

class AdminSizeStandardViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]

    def get_permissions(self):
        if self.action in ['create_size', 'update_size', 'delete_size']:
            return [IsAuthenticated(), IsAdminUserPermission()]
        return [IsAuthenticated()]
    @action(detail=False, methods=['get'])
    def get_all_sizes(self, request):
        from .models import ClothingSizeGOST, FootwearSizeGOST, HeadwearSizeGOST

        size_type = request.query_params.get('size_type')

        if size_type == 'clothing':
            sizes = ClothingSizeGOST.objects.all().order_by('sort_order')
            result = [{
                'size_id': s.size_id,
                'gender': s.gender,
                'size_code': s.size_code,
                'size_name': s.size_name,
                'height_min': s.height_min,
                'height_max': s.height_max,
                'chest_circumference': s.chest_circumference,
                'waist_circumference': s.waist_circumference,
                'hip_circumference': s.hip_circumference,
                'sort_order': s.sort_order,
            } for s in sizes]
        elif size_type == 'footwear':
            sizes = FootwearSizeGOST.objects.all().order_by('sort_order')
            result = [{
                'size_id': s.size_id,
                'size_ru': s.size_ru,
                'size_eu': s.size_eu,
                'size_us': s.size_us,
                'foot_length_min': s.foot_length_min,
                'foot_length_max': s.foot_length_max,
                'sort_order': s.sort_order,
            } for s in sizes]
        elif size_type == 'headwear':
            sizes = HeadwearSizeGOST.objects.all().order_by('sort_order')
            result = [{
                'size_id': s.size_id,
                'size_code': s.size_code,
                'head_circumference_min': s.head_circumference_min,
                'head_circumference_max': s.head_circumference_max,
                'sort_order': s.sort_order,
            } for s in sizes]
        else:
            return Response({'error': 'Invalid size_type'}, status=400)

        return Response(result)

    @action(detail=False, methods=['get'])
    def auto_select_size(self, request):
        """Автоматически подобрать размер на основе данных сотрудника"""
        employee_id = request.query_params.get('employee_id')
        nomenclature_id = request.query_params.get('nomenclature_id')

        if not employee_id:
            return Response({'error': 'employee_id required'}, status=400)

        try:
            employee = Employee.objects.get(employee_id=employee_id)
            nomenclature = Nomenclature.objects.get(nomenclature_id=nomenclature_id)

            title = nomenclature.title.lower()

            # Определяем тип СИЗ
            if 'обув' in title or 'сапог' in title or 'ботин' in title:
                size_type = 'footwear'
                suggested_size = get_footwear_size(employee)
            elif 'каск' in title or 'шлем' in title or 'шапк' in title:
                size_type = 'headwear'
                suggested_size = get_headwear_size(employee)
            else:
                size_type = 'clothing'
                suggested_size = get_clothing_size(employee)

            return Response({
                'suggested_size': suggested_size,
                'size_type': size_type,
                'employee_height': employee.heightcm,
                'employee_gender': employee.gender,
            })
        except Employee.DoesNotExist:
            return Response({'error': 'Employee not found'}, status=404)
        except Nomenclature.DoesNotExist:
            return Response({'error': 'Nomenclature not found'}, status=404)

    def get_clothing_size(employee):
        """Подобрать размер одежды по ГОСТ на основе роста"""
        from .models import ClothingSizeGOST

        if not employee.heightcm:
            return None

        # Определяем пол сотрудника
        gender = employee.gender if employee.gender else 'M'

        # Ищем подходящий размер по росту
        sizes = ClothingSizeGOST.objects.filter(gender=gender).order_by('sort_order')

        for size in sizes:
            height_min = size.height_min or 0
            height_max = size.height_max or 300
            if height_min <= employee.heightcm <= height_max:
                return size.size_code

        # Если не нашли, возвращаем средний размер
        if sizes.exists():
            middle_index = sizes.count() // 2
            return list(sizes)[middle_index].size_code

        return None
    def get_footwear_size(employee):
        """Подобрать размер обуви по ГОСТ (если есть длина стопы)"""
        from .models import FootwearSizeGOST

        # Если есть сохраненный размер обуви, используем его
        if employee.shoesize:
            return str(employee.shoesize)

        # Иначе - стандартный размер по умолчанию
        sizes = FootwearSizeGOST.objects.all().order_by('sort_order')
        if sizes.exists():
            # По умолчанию 42 размер (мужской) или 38 (женский)
            if employee.gender == 'F':
                return str(sizes.filter(size_ru=38).first().size_ru if sizes.filter(size_ru=38).exists() else '38')
            else:
                return str(sizes.filter(size_ru=42).first().size_ru if sizes.filter(size_ru=42).exists() else '42')

        return None

    def get_headwear_size(employee):
        """Подобрать размер головного убора по ГОСТ"""
        from .models import HeadwearSizeGOST

        if employee.headsize:
            return str(employee.headsize)

        # Стандартный размер по умолчанию
        sizes = HeadwearSizeGOST.objects.all().order_by('sort_order')
        if sizes.exists():
            default_size = sizes.filter(size_code='58').first()
            if default_size:
                return default_size.size_code

        return None
    @action(detail=False, methods=['get'], url_path='by-type-gender')
    def get_sizes_by_type_gender(self, request):
        """Получить размеры по типу и полу"""
        from .models import ClothingSizeGOST, FootwearSizeGOST, HeadwearSizeGOST

        size_type = request.query_params.get('size_type')
        gender = request.query_params.get('gender')  # M, F

        if size_type == 'clothing':
            sizes = ClothingSizeGOST.objects.all().order_by('sort_order')
            if gender:
                sizes = sizes.filter(gender=gender)
            result = [{
                'size_id': s.size_id,
                'size_code': s.size_code,
                'size_name': s.size_name,
                'height_min': s.height_min,
                'height_max': s.height_max,
                'chest_circumference': s.chest_circumference,
            } for s in sizes]
        elif size_type == 'footwear':
            sizes = FootwearSizeGOST.objects.all().order_by('sort_order')
            result = [{
                'size_id': s.size_id,
                'size_ru': s.size_ru,
                'size_eu': s.size_eu,
                'size_us': s.size_us,
                'foot_length_min': s.foot_length_min,
                'foot_length_max': s.foot_length_max,
            } for s in sizes]
        elif size_type == 'headwear':
            sizes = HeadwearSizeGOST.objects.all().order_by('sort_order')
            result = [{
                'size_id': s.size_id,
                'size_code': s.size_code,
                'head_circumference_min': s.head_circumference_min,
                'head_circumference_max': s.head_circumference_max,
            } for s in sizes]
        else:
            return Response({'error': 'Invalid size_type'}, status=400)

        return Response(result)
    @action(detail=False, methods=['post'])
    def create_size(self, request):
        from .models import ClothingSizeGOST, FootwearSizeGOST, HeadwearSizeGOST

        size_type = request.data.get('size_type')
        data = request.data

        try:
            if size_type == 'clothing':
                size = ClothingSizeGOST.objects.create(
                    gender=data.get('gender'),
                    size_code=data.get('size_code'),
                    size_name=data.get('size_name', ''),
                    height_min=data.get('height_min'),
                    height_max=data.get('height_max'),
                    chest_circumference=data.get('chest_circumference'),
                    waist_circumference=data.get('waist_circumference'),
                    hip_circumference=data.get('hip_circumference'),
                    sort_order=data.get('sort_order'),
                )
            elif size_type == 'footwear':
                size = FootwearSizeGOST.objects.create(
                    size_ru=data.get('size_ru'),
                    size_eu=data.get('size_eu'),
                    size_us=data.get('size_us'),
                    foot_length_min=data.get('foot_length_min'),
                    foot_length_max=data.get('foot_length_max'),
                    sort_order=data.get('sort_order'),
                )
            elif size_type == 'headwear':
                size = HeadwearSizeGOST.objects.create(
                    size_code=data.get('size_code'),
                    head_circumference_min=data.get('head_circumference_min'),
                    head_circumference_max=data.get('head_circumference_max'),
                    sort_order=data.get('sort_order'),
                )
            else:
                return Response({'error': 'Invalid size_type'}, status=400)

            return Response({'message': 'Size created', 'size_id': size.size_id})
        except Exception as e:
            return Response({'error': str(e)}, status=400)

    @action(detail=False, methods=['put'])
    def update_size(self, request):
        from .models import ClothingSizeGOST, FootwearSizeGOST, HeadwearSizeGOST

        size_id = request.data.get('size_id')
        size_type = request.data.get('size_type')
        data = request.data

        try:
            if size_type == 'clothing':
                size = ClothingSizeGOST.objects.get(size_id=size_id)
                if data.get('gender'): size.gender = data['gender']
                if data.get('size_code'): size.size_code = data['size_code']
                if data.get('size_name'): size.size_name = data['size_name']
                if data.get('height_min') is not None: size.height_min = data['height_min']
                if data.get('height_max') is not None: size.height_max = data['height_max']
                if data.get('chest_circumference') is not None: size.chest_circumference = data['chest_circumference']
                if data.get('waist_circumference') is not None: size.waist_circumference = data['waist_circumference']
                if data.get('hip_circumference') is not None: size.hip_circumference = data['hip_circumference']
                if data.get('sort_order') is not None: size.sort_order = data['sort_order']
                size.save()
            elif size_type == 'footwear':
                size = FootwearSizeGOST.objects.get(size_id=size_id)
                if data.get('size_ru'): size.size_ru = data['size_ru']
                if data.get('size_eu'): size.size_eu = data['size_eu']
                if data.get('size_us'): size.size_us = data['size_us']
                if data.get('foot_length_min') is not None: size.foot_length_min = data['foot_length_min']
                if data.get('foot_length_max') is not None: size.foot_length_max = data['foot_length_max']
                if data.get('sort_order') is not None: size.sort_order = data['sort_order']
                size.save()
            elif size_type == 'headwear':
                size = HeadwearSizeGOST.objects.get(size_id=size_id)
                if data.get('size_code'): size.size_code = data['size_code']
                if data.get('head_circumference_min') is not None: size.head_circumference_min = data['head_circumference_min']
                if data.get('head_circumference_max') is not None: size.head_circumference_max = data['head_circumference_max']
                if data.get('sort_order') is not None: size.sort_order = data['sort_order']
                size.save()
            else:
                return Response({'error': 'Invalid size_type'}, status=400)

            return Response({'message': 'Size updated'})
        except Exception as e:
            return Response({'error': str(e)}, status=404)

    @action(detail=False, methods=['delete'])
    def delete_size(self, request):
        from .models import ClothingSizeGOST, FootwearSizeGOST, HeadwearSizeGOST

        size_id = request.data.get('size_id')
        size_type = request.data.get('size_type')

        try:
            if size_type == 'clothing':
                ClothingSizeGOST.objects.get(size_id=size_id).delete()
            elif size_type == 'footwear':
                FootwearSizeGOST.objects.get(size_id=size_id).delete()
            elif size_type == 'headwear':
                HeadwearSizeGOST.objects.get(size_id=size_id).delete()
            else:
                return Response({'error': 'Invalid size_type'}, status=400)

            return Response({'message': 'Size deleted'})
        except Exception as e:
            return Response({'error': str(e)}, status=404)

class NomenclatureViewSet(viewsets.ReadOnlyModelViewSet):
    """API для получения списка СИЗ (номенклатуры)"""
    permission_classes = [IsAuthenticated]
    queryset = Nomenclature.objects.filter(is_active=True)
    serializer_class = NomenclatureSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['title']
class EmployeeViewSet(viewsets.ViewSet):
    """API для работы с сотрудниками"""
    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=['get'])
    def my_shop_employees(self, request):
        """Получить сотрудников цеха текущего пользователя"""
        try:
            app_user = AppUser.objects.get(login=request.user.username)
            employees = Employee.objects.filter(
                shop=app_user.shop,
                is_active=True
            ).select_related('position')

            result = []
            for emp in employees:
                result.append({
                    'employee_id': emp.employee_id,
                    'full_name': emp.full_name,
                    'position_name': emp.position.title,
                    'gender': emp.gender,
                    'heightcm': emp.heightcm,
                    'clothing_size': emp.clothing_size,
                    'shoesize': emp.shoesize,
                    'headsize': emp.headsize,
                })
            return Response(result)
        except AppUser.DoesNotExist:
            return Response([])

    @action(detail=False, methods=['get'])
    def ppe_history(self, request):
        """Получить историю выдач СИЗ для сотрудника"""
        employee_id = request.query_params.get('employee_id')
        if not employee_id:
            return Response({'error': 'employee_id required'}, status=400)

        try:
            issues = PPEIssue.objects.filter(
                employee_id=employee_id
            ).select_related('nomenclature').order_by('-issue_date')

            result = []
            for issue in issues:
                result.append({
                    'issue_id': issue.issue_id,
                    'issue_date': issue.issue_date,
                    'nomenclature_title': issue.nomenclature.title,
                    'size': issue.size,
                    'quantity': float(issue.quantity),
                    'next_issue_date': issue.next_issue_date,
                    'unit': issue.nomenclature.unit,
                })
            return Response(result)
        except Exception as e:
            return Response({'error': str(e)}, status=400)

    @action(detail=True, methods=['get'])
    def history(self, request, pk=None):
        """Получить историю антропометрии сотрудника"""
        try:
            employee = Employee.objects.get(employee_id=pk)
            # Здесь нужно будет добавить модель истории антропометрии
            return Response([])
        except Employee.DoesNotExist:
            return Response({'error': 'Employee not found'}, status=404)

    @action(detail=True, methods=['patch'])
    def update_anthropometry(self, request, pk=None):
        """Обновить антропометрические данные сотрудника"""
        try:
            employee = Employee.objects.get(employee_id=pk)

            # Проверяем права (только начальник цеха или админ)
            app_user = AppUser.objects.get(login=request.user.username)
            if not (request.user.is_superuser or app_user.shop == employee.shop):
                return Response({'error': 'No permission'}, status=403)

            if 'heightcm' in request.data:
                employee.heightcm = request.data['heightcm']
            if 'clothing_size' in request.data:
                employee.clothing_size = request.data['clothing_size']
            if 'shoesize' in request.data:
                employee.shoesize = request.data['shoesize']
            if 'headsize' in request.data:
                employee.headsize = request.data['headsize']
            if 'gender' in request.data:
                employee.gender = request.data['gender']

            employee.save()
            return Response({'message': 'Employee updated successfully'})
        except Employee.DoesNotExist:
            return Response({'error': 'Employee not found'}, status=404)
        except AppUser.DoesNotExist:
            return Response({'error': 'User not found'}, status=404)
